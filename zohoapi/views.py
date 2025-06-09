from rest_framework.decorators import api_view, permission_classes
import requests
import random
import string
import uuid
from babel.numbers import get_currency_symbol
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from rest_framework.exceptions import ValidationError
from operationsBackend import settings
from django.utils.crypto import get_random_string
from rest_framework.exceptions import AuthenticationFailed
from rest_framework import generics
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import generics, filters
from django.contrib.auth.models import User
from rest_framework.permissions import AllowAny, IsAuthenticated
import pandas as pd
from django.db.models import Sum, Max
from django.db.models.functions import Coalesce
from api.utils.methods import get_subordinates_of_a_user_in_role
from api.models import Role, Sales, Profile
from num2words import num2words
from datetime import timedelta
from api.serializers import get_exchange_rate, GmSheetSerializer
from openpyxl import Workbook
import json
from rest_framework.views import APIView

from django.db.models import Q, Subquery, OuterRef, Prefetch, Value, Sum, F
from django.db.models.functions import Coalesce
from django.contrib.auth import authenticate, login, logout
from django.core.mail import send_mail
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie
from rest_framework import status
from django.http import HttpResponse
from .serializers import (
    InvoiceDataEditSerializer,
    InvoiceDataSerializer,
    VendorSerializer,
    InvoiceStatusUpdateGetSerializer,
    VendorEditSerializer,
    ZohoVendorSerializer,
    SalesOrderGetSerializer,
    PurchaseOrderSerializer,
    PurchaseOrderGetSerializer,
    ClientInvoiceSerializer,
    ClientInvoiceGetSerializer,
    InvoiceDataGetSerializer,
    SalesOrderSerializer,
    ContactSerializer,
    DealSerializer,
    DealSerializerDepthOne,
    CompanySerializer,
    EntitySerializer,
    ZohoCustomerSerializer,
    SalesOrderLineItemSerializer,
    V2SalesOrderLineItemSerializer,
    V2ClientInvoiceLineItemSerializer,
    ClientInvoiceLineItemSerializer,
    BankDetailsSerializer,
    PaymentSerializer,
    PaymentCreateSerializer,
    SalesOrderLineItemListSerializer,
    FacilitatorMonthlyPOSerializer,
    CreditNoteSerializer,
    PurchaseOrderLineItemSerializer,
    V2BillSerializer,
    OPEPurchaseOrderSerializer,
)

from zohoapi.utils.methods import (
    fetch_and_filter_purchase_orders,
    fetch_and_process_invoices,
    get_line_items_details,
    fetch_invoices_db,
    get_purchase_order_ids_for_project,
    get_current_financial_year,
    generate_new_po_number,
    generate_new_ctt_po_number,
    get_current_month_start_and_end_date,
    generate_new_invoice_number,
    generate_new_so_number,
    get_current_financial_year_dates,
    create_payload_data,
    get_revenue_data,
    calculate_financials_from_orders,
    get_meeraq_sales_orders,
    zoho_api_request,
    create_purchase_order,
    calculate_total_revenue,
    calculate_total_cost,
    create_or_update_contacts,
    create_or_update_company,
    create_or_update_deals,
    get_owner_details,
    process_line_item_custom_fields,
    get_purchase_order_totals,
    get_purchase_order_instance_totals,
    process_po_line_item_data,
    create_custom_field_data,
    create_singapore_purchase_order,
    filter_invoice_data,
    map_bill_to_invoice,
)
from zohoapi.utils.common import (
    get_financial_year,
    get_invoice_data_for_pdf,
    amount_convert_to_words,
    salesorder_line_item_due_date_case,
    get_styles,
    update_invoice_status_and_balance,
)
from api.utils.auth import get_role_response
from api.utils.email import send_mail_templates, send_mail_templates_with_attachment
from api.utils.pagination import CustomPageNumberPagination
from .models import (
    InvoiceData,
    Vendor,
    InvoiceStatusUpdate,
    ZohoVendor,
    PurchaseOrder,
    SalesOrder,
    Bill,
    ClientInvoice,
    SalesOrderLineItem,
    ClientInvoiceLineItem,
    Deal,
    Contact,
    Company,
    Entity,
    ZohoCustomer,
    BankDetails,
    Payment,
    BillLineItem,
    CreditNote,
    PurchaseOrderLineItem,
    GmSheet,
    HandoverDetails,
)
import base64
from io import BytesIO
import environ
import os
from django.http import HttpResponse
from django.http import JsonResponse
import pdfkit
from django.middleware.csrf import get_token
from django.db import transaction
from django.core.exceptions import ObjectDoesNotExist
from collections import defaultdict
import io
from datetime import datetime
import xlsxwriter
from decimal import Decimal
from collections import defaultdict
from api.permissions import IsInRoles
from zohoapi.utils.auth import get_user_data
from zohoapi.utils.common import SerializerByMethodMixin
from .filters import (
    PaymentFilters,
    SalesOrderLineItemFilter,
    CompanyFilter,
    OpePurchaseOrderFilter,
    ContactFilter,
)
from django.views.decorators.csrf import csrf_exempt

env = environ.Env()

wkhtmltopdf_path = os.environ.get(
    "WKHTMLTOPDF_PATH", r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"
)

pdfkit_config = pdfkit.configuration(wkhtmltopdf=f"{wkhtmltopdf_path}")


def get_paid_so_item_ids(sales_order):
    ci_line_items = ClientInvoiceLineItem.objects.filter(
        clientinvoice__isnull=False,
        clientinvoice__status="paid",
        clientinvoice__salesorder_id=sales_order.salesorder_id,
    )

    quantity_mapping = {}
    for item in sales_order.so_line_items.all():
        quantity_mapping[item.line_item_id] = item.quantity

    line_item_quantity_paid_mapping = {}
    for item in ci_line_items:
        if item.salesorder_item_id in line_item_quantity_paid_mapping:
            line_item_quantity_paid_mapping[item.salesorder_item_id] += item.quantity
        else:
            line_item_quantity_paid_mapping[item.salesorder_item_id] = item.quantity

    completely_paid_line_item_ids = []
    for key, value in line_item_quantity_paid_mapping.items():
        if value == quantity_mapping[key]:
            completely_paid_line_item_ids.append(key)

    return completely_paid_line_item_ids


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("vendor", "pmo", "finance")])
def get_purchase_orders(request, vendor_id):
    data, http_status = fetch_and_filter_purchase_orders(vendor_id)
    return Response(data, status=http_status)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_total_revenue(request, vendor_id):
    try:
        invoices = InvoiceData.objects.filter(vendor_id=vendor_id).order_by(
            "-created_at"
        )
        currency_symbol = ""
        if invoices.first():
            currency_symbol = invoices.first().currency_symbol

        invoices = filter_invoice_data(invoices)
        # Initialize with Decimal(0) instead of float 0.0
        total_revenue = Decimal("0.0")
        for invoice in invoices:
            total_revenue += invoice.total
        # Convert to float at the end if needed
        return Response(
            {"total_revenue": total_revenue, "currency_symbol": currency_symbol}
        )
    except Exception as e:
        return Response({"error": str(e)}, status=400)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("vendor", "pmo", "finance")])
def get_invoices_with_status(request, vendor_id, purchase_order_id):
    invoice_res, http_status = fetch_and_process_invoices(vendor_id, purchase_order_id)
    return Response(invoice_res, status=http_status)


def get_paid_po_item_ids(purchase_order):
    paid_po_item_ids = []

    for line_item in purchase_order["line_items"]:
        po_item_id = line_item["line_item_id"]
        po_item_quantity = Decimal(str(line_item["quantity"]))

        # Get all paid bill line items for this PO item
        billed_quantity = BillLineItem.objects.filter(
            bill__status="paid", purchaseorder_item_id=po_item_id
        ).aggregate(total_quantity=Coalesce(Sum("quantity"), Decimal("0")))[
            "total_quantity"
        ]

        # Compare with a small tolerance to handle floating point precision
        if abs(po_item_quantity - billed_quantity) < Decimal("0.000001"):
            paid_po_item_ids.append(po_item_id)

    return paid_po_item_ids


def get_paid_po_item_ids_instance(purchase_order):
    """
    Get line item IDs from a purchase order that have been fully paid through bills.

    Args:
        purchase_order: PurchaseOrder instance to check paid items for

    Returns:
        list: List of line item IDs that have been fully paid

    Note:
        Uses a small tolerance (0.000001) for decimal comparisons to handle floating point precision
    """

    paid_po_item_ids = []
    DECIMAL_TOLERANCE = Decimal("0.000001")

    try:
        # Fetch all line items in a single query
        po_line_items = purchase_order.po_line_items.all().select_related()

        # Get all billed quantities in a single query
        billed_quantities = (
            BillLineItem.objects.filter(
                bill__status="paid",
                purchaseorder_item_id__in=[item.line_item_id for item in po_line_items],
            )
            .values("purchaseorder_item_id")
            .annotate(total_quantity=Coalesce(Sum("quantity"), Decimal("0")))
        )

        # Convert to dictionary for O(1) lookup
        billed_qty_dict = {
            item["purchaseorder_item_id"]: item["total_quantity"]
            for item in billed_quantities
        }

        for line_item in po_line_items:
            try:
                po_item_quantity = Decimal(str(line_item.quantity))
                billed_quantity = billed_qty_dict.get(
                    line_item.line_item_id, Decimal("0")
                )

                print(
                    f"DEBUG: Line item {line_item.line_item_id} - "
                    f"PO Qty: {po_item_quantity}, Billed Qty: {billed_quantity}"
                )

                if abs(po_item_quantity - billed_quantity) < DECIMAL_TOLERANCE:
                    paid_po_item_ids.append(line_item.line_item_id)

            except (TypeError, ValueError) as e:
                print(
                    f"ERROR: Failed to process line item {line_item.line_item_id}: {str(e)}"
                )
                continue

        return paid_po_item_ids

    except Exception as e:
        print(f"ERROR: Failed to get paid PO items: {str(e)}")
        return []


def get_purchase_order_data_util(purchaseorder_id):
    try:
        purchase_order_data = PurchaseOrder.objects.filter(
            purchaseorder_id=purchaseorder_id
        ).first()
        if not purchase_order_data:
            return None, status.HTTP_404_NOT_FOUND, "Purchase order not found"

        print("step 2")
        # For non-India entity
        serializer = PurchaseOrderSerializer(purchase_order_data)
        line_items = PurchaseOrderLineItemSerializer(
            purchase_order_data.po_line_items.all(), many=True
        ).data
        print("step 3")
        paid_so_line_item_ids = get_paid_po_item_ids_instance(purchase_order_data)
        print("step 5")
        complete_data = {
            **serializer.data,
            "paid_po_item_ids": paid_so_line_item_ids,
            "line_items": line_items,
        }
        print("step 4")

        return complete_data, status.HTTP_200_OK, None
    except Exception as e:
        print(str(e))


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("vendor", "pmo", "finance")])
def get_purchase_order_data(request, purchaseorder_id):
    try:
        purchase_order = PurchaseOrder.objects.filter(
            purchaseorder_id=purchaseorder_id
        ).first()
        if not purchase_order:
            return Response(
                {"error": "Purchase order not found"}, status=status.HTTP_404_NOT_FOUND
            )

        data, status_code, error = get_purchase_order_data_util(purchaseorder_id)
        if error:
            return Response({"error": error}, status=status_code)
        return Response(data, status=status_code)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("vendor")])
def add_invoice_data(request):
    try:
        with transaction.atomic():
            invoice_date = datetime.strptime(
                request.data["invoice_date"], "%Y-%m-%d"
            ).date()
            start_year, end_year = get_financial_year(invoice_date)
            invoices = InvoiceData.objects.filter(
                vendor_id=request.data["vendor_id"],
                invoice_number=request.data["invoice_number"],
                invoice_date__range=(f"{start_year}-04-01", f"{end_year}-03-31"),
            )
            hsn_or_sac = request.data.get("hsn_or_sac", None)
            vendor = Vendor.objects.get(vendor_id=request.data["vendor_id"])
            if hsn_or_sac:
                try:
                    print(
                        f"Updating vendor HSN/SAC for vendor_id: {request.data['vendor_id']}"
                    )

                    vendor.hsn_or_sac = hsn_or_sac
                    vendor.save()
                    print("Vendor HSN/SAC updated successfully")
                except Exception as e:
                    print(f"Error updating vendor HSN/SAC: {str(e)}")
                    return Response({"error": "Vendor not found."}, status=400)

            if invoices.count() > 0:
                return Response(
                    {"error": "Invoice number should be unique."}, status=400
                )

            print("Fetching vendor details")

            print("Validating serializer data")
            serializer = InvoiceDataSerializer(data=request.data)
            vendor_details = vendor.zoho_vendor
            if serializer.is_valid():
                print("Serializer is valid, saving invoice")
                invoice_instance = serializer.save()

                print("Getting vendor details")

                invoice_instance.vendor_name = vendor_details.contact_name
                # updating exchange rate and the purchase order reference
                if invoice_instance.purchase_order_id:
                    print(
                        f"Fetching purchase order for ID: {invoice_instance.purchase_order_id}"
                    )
                    try:
                        purchase_order = PurchaseOrder.objects.get(
                            purchaseorder_id=invoice_instance.purchase_order_id
                        )
                        # Set the purchase order reference
                        invoice_instance.purchase_order_reference = purchase_order
                        # Set the exchange rate if available
                        if (
                            hasattr(purchase_order, "exchange_rate")
                            and purchase_order.exchange_rate
                        ):
                            invoice_instance.exchange_rate = (
                                purchase_order.exchange_rate
                            )
                            print(f"Exchange rate set: {purchase_order.exchange_rate}")
                    except PurchaseOrder.DoesNotExist:
                        print(
                            f"Purchase order not found for ID: {invoice_instance.purchase_order_id}"
                        )
                    except Exception as e:
                        print(f"Error linking purchase order: {str(e)}")
                invoice_instance.save()
                print(
                    f"Updated invoice with vendor name: {vendor_details.contact_name}"
                )

                approver_email = serializer.data["approver_email"]
                print(f"Approver email: {approver_email}")

                print("Preparing invoice data for PDF")
                invoice_data = get_invoice_data_for_pdf(
                    InvoiceDataSerializer(invoice_instance).data, vendor.hsn_or_sac
                )

                print("Sending email notification")
                email_recipient = (
                    approver_email
                    if env("ENVIRONMENT") == "PRODUCTION"
                    else "tech@meeraq.com"
                )

                send_mail_templates(
                    "vendors/add_invoice_sg.html",
                    [email_recipient],
                    f"{vendor_details.entity.country} Action Needed: Approval Required for Invoice - {invoice_data['vendor_name']} -> {invoice_data['invoice_number']}",
                    {**invoice_data, "link": env("CAAS_APP_URL")},
                    [],
                )

                print("Email sent successfully")

                return Response(
                    {"message": "Invoice generated successfully"}, status=201
                )
            else:
                print(f"Serializer validation errors: {serializer.errors}")
                return Response(serializer.errors, status=400)
    except Exception as e:
        print(f"Critical error in add_invoice_data: {str(e)}")
        print(f"Error type: {type(e)}")
        return Response({"error": "Failed to generate invoice"}, status=500)


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("vendor")])
def edit_invoice(request, invoice_id):
    invoice = get_object_or_404(InvoiceData, id=invoice_id)
    if (
        InvoiceData.objects.filter(
            vendor_id=invoice.vendor_id, invoice_number=request.data["invoice_number"]
        )
        .exclude(id=invoice.id)
        .exists()
    ):
        return Response({"error": "Invoice already exists with the invoice number"})
    serializer = InvoiceDataEditSerializer(data=request.data, instance=invoice)
    vendor = Vendor.objects.get(vendor_id=request.data["vendor_id"])

    if serializer.is_valid():
        serializer.save()
        invoice.status = "in_review"
        invoice.save()
        approver_email = invoice.approver_email
        invoice_data_serialized = InvoiceDataSerializer(invoice).data
        hsn_or_sac = vendor.hsn_or_sac
        invoice_data = get_invoice_data_for_pdf(invoice_data_serialized, hsn_or_sac)
        # invoice_data = get_invoice_data_for_pdf(InvoiceDataSerializer(invoice).data, vendor.hsn_or_sac)

        send_mail_templates(
            "vendors/edit_invoice.html",
            [
                (
                    approver_email
                    if env("ENVIRONMENT") == "PRODUCTION"
                    else "tech@meeraq.com"
                )
            ],
            f"Action Needed: Re-Approval Required for Invoice - {invoice_data['vendor_name']}  + {invoice_data['invoice_number']}",
            {**invoice_data, "link": env("CAAS_APP_URL")},
            [],
        )
        return Response({"message": "Invoice edited successfully."}, status=201)
    else:
        return Response({"error": "Invalid data."}, status=400)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated, IsInRoles("vendor", "finance", "pmo")])
def delete_invoice(request, invoice_id):
    try:
        invoice = get_object_or_404(InvoiceData, id=invoice_id)
        invoice.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    except InvoiceData.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("vendor", "pmo", "finance")])
def get_purchase_order_and_invoices(request, purchase_order_id):
    try:
        purchase_order_data = PurchaseOrder.objects.filter(
            purchaseorder_id=purchase_order_id
        ).first()
        if not purchase_order_data:
            return Response(
                {"error": "Purchase order not found"}, status=status.HTTP_404_NOT_FOUND
            )

        purchase_orders_serializer = PurchaseOrderSerializer(purchase_order_data).data
        invoices = InvoiceData.objects.filter(purchase_order_id=purchase_order_id)
        invoice_serializer = InvoiceDataSerializer(invoices, many=True)

        return Response(
            {
                "purchase_order": purchase_orders_serializer,
                "invoices": invoice_serializer.data,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to fetch purchase order data"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([AllowAny])
def get_vendor_exists_and_not_existing_emails(request):
    access_token = get_access_token(env("ZOHO_REFRESH_TOKEN"))
    if access_token:
        headers = {"Authorization": f"Bearer {access_token}"}
        url = f"{base_url}/contacts?organization_id={env('ZOHO_ORGANIZATION_ID')}&contact_type=vendor"
        vendor_response = requests.get(url, headers=headers)
        if vendor_response.json()["message"] == "success":
            existing_vendors = []
            not_existing_vendors = []
            for vendor in vendor_response.json()["contacts"]:
                if vendor["email"]:
                    try:
                        existing_vendor = Vendor.objects.get(email=vendor["email"])
                        existing_vendors.append(vendor["email"])
                    except Vendor.DoesNotExist:
                        not_existing_vendors.append(vendor["email"])
                        pass
            return Response(
                {
                    "vendors": vendor_response.json()["contacts"],
                    "existing_vendors": existing_vendors,
                    "not_existing_vendors": not_existing_vendors,
                },
                status=200,
            )
        else:
            return Response({"error": "Failed to get vendors."}, status=400)
    else:
        return Response({"error": "Unauthorized	."}, status=400)


@api_view(["GET"])
@permission_classes([AllowAny])
def import_invoices_from_zoho(request):
    access_token = get_access_token(env("ZOHO_REFRESH_TOKEN"))
    if access_token:
        headers = {"Authorization": f"Bearer {access_token}"}
        vendors = Vendor.objects.all()
        res = []
        bill_details_res = []
        for vendor in vendors:
            import_invoices_for_vendor_from_zoho(vendor, headers, res, bill_details_res)
        return Response({"res": res, "bill_details_res": bill_details_res}, status=200)
    else:
        return Response({"error": "Invalid invoices"}, status=400)


@api_view(["GET"])
@permission_classes([AllowAny])
def export_invoice_data(request):
    # Retrieve all InvoiceData objects
    queryset = InvoiceData.objects.all()

    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers to the worksheet
    headers = [
        "Vendor ID",
        "Vendor Name",
        "Vendor Email",
        "Vendor Billing Address",
        "Vendor GST",
        "Vendor Phone",
        "Purchase Order ID",
        "Purchase Order No",
        "Invoice Number",
        "Customer Name",
        "Customer Notes",
        "Customer GST",
        "Total",
        "Is Oversea Account",
        "TIN Number",
        "Type of Code",
        "IBAN",
        "SWIFT Code",
        "Invoice Date",
        "Beneficiary Name",
        "Bank Name",
        "Account Number",
        "IFSC Code",
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data to the worksheet
    for row_num, invoice_data in enumerate(queryset, 2):
        ws.append(
            [
                invoice_data.vendor_id,
                invoice_data.vendor_name,
                invoice_data.vendor_email,
                invoice_data.vendor_billing_address,
                invoice_data.vendor_gst,
                invoice_data.vendor_phone,
                invoice_data.purchase_order_id,
                invoice_data.purchase_order_no,
                invoice_data.invoice_number,
                invoice_data.customer_name,
                invoice_data.customer_notes,
                invoice_data.customer_gst,
                invoice_data.total,
                invoice_data.is_oversea_account,
                invoice_data.tin_number,
                invoice_data.type_of_code,
                invoice_data.iban,
                invoice_data.swift_code,
                invoice_data.invoice_date,
                invoice_data.beneficiary_name,
                invoice_data.bank_name,
                invoice_data.account_number,
                invoice_data.ifsc_code,
            ]
        )

    # Create a response with the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=invoice_data.xlsx"
    wb.save(response)

    return response


class DownloadInvoice(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo", "vendor", "finance")]

    def get(self, request, record_id):
        try:
            invoice = get_object_or_404(InvoiceData, id=record_id)
            vendor = Vendor.objects.get(vendor_id=invoice.vendor_id)
            invoice_data = get_invoice_data_for_pdf(
                InvoiceDataSerializer(invoice).data, vendor.hsn_or_sac
            )

            image_base64 = None
            try:
                image_url = f"{invoice_data['signature']}"
                # Attempt to send the email
                image_response = requests.get(image_url)
                image_response.raise_for_status()
                # Convert the downloaded image to base64
                image_base64 = base64.b64encode(image_response.content).decode("utf-8")
            except Exception as e:
                pass
            email_message = render_to_string(
                "invoice_pdf.html",
                {"invoice": invoice_data, "image_base64": image_base64},
            )
            pdf = pdfkit.from_string(email_message, False, configuration=pdfkit_config)
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = (
                f'attachment; filename={f"{invoice.invoice_number}_invoice.pdf"}'
            )
            return response

        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to download invoice."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


def format_address(input_address):
    attention = input_address.get("attention")
    address = input_address.get("address")
    street2 = input_address.get("street2")
    city = input_address.get("city")
    state = input_address.get("state")
    zip_code = input_address.get("zip")
    country = input_address.get("country")

    address_details = [attention, address, street2, city, state, zip_code, country]
    formatted_address = ", ".join(filter(None, address_details))

    return formatted_address


def get_Client_invoice_line_items(line_items):
    data = []
    for line_item in line_items:
        line_item_data = {
            "id": line_item.id,
            "name": line_item.description,
            "quantity": line_item.quantity,
            "rate": line_item.rate,
            "amount": line_item.quantity * line_item.rate,
            "quantity_mul_rate": round(line_item.quantity * line_item.rate, 2),
            "hsn_or_sac": line_item.hsn_or_sac,
        }
        data.append(line_item_data)

    return data


def format_key(key):
    return key.replace("_", " ").title()


class DownloadClientInvoice(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo", "vendor", "finance")]

    def get(self, request, record_id):
        try:
            invoice = get_object_or_404(ClientInvoice, id=record_id)
            bank_detail = (
                BankDetailsSerializer(invoice.bank_detail).data
                if invoice.bank_detail
                else {}
            )
            bank_detail = {
                format_key(key): value
                for key, value in bank_detail.items()
                if value and key not in ["id", "display_name"]
            }
            has_bank_details = any(value for value in bank_detail.values())
            amount_to_words = amount_convert_to_words(
                Decimal(invoice.total), invoice.currency_code
            )
            data = {
                "organization_name": invoice.sales_order.entity.name,
                "organization_billing_address": invoice.sales_order.entity.billing_address,
                "customer_name": invoice.sales_order.zoho_customer.contact_name,
                "customer_address": format_address(
                    invoice.sales_order.zoho_customer.billing_address
                ),
                "customer_email": invoice.sales_order.zoho_customer.email,
                "customer_phone": invoice.sales_order.zoho_customer.phone,
                "invoice_number": invoice.invoice_number,
                "customer_notes": invoice.notes,
                "invoice_date": (
                    invoice.date.strftime("%d/%m/%Y")
                    if invoice.date
                    else None if invoice.date else ""
                ),
                "due_date": (
                    invoice.due_date.strftime("%d/%m/%Y")
                    if invoice.due_date
                    else None if invoice.due_date else ""
                ),
                "sales_order_id": invoice.sales_order.salesorder_id,
                "sales_order_no": invoice.sales_order.salesorder_number,
                "currency_code": invoice.currency_code,
                "currency_symbol": invoice.currency_symbol,
                "line_items": get_Client_invoice_line_items(
                    invoice.client_invoice_line_items.all()
                ),
                "total": invoice.total,
                "bank_detail": bank_detail,
                "has_bank_details": has_bank_details,
                "amount_to_words": amount_to_words,
            }

            email_message = render_to_string(
                "client_invoice.html",
                data,
            )
            pdf = pdfkit.from_string(email_message, False, configuration=pdfkit_config)
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = (
                f'attachment; filename={f"{invoice.invoice_number}_invoice.pdf"}'
            )
            return response

        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to download invoice."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class DownloadAttatchedInvoice(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo", "vendor", "finance")]

    def get(self, request, record_id):
        try:
            invoice = get_object_or_404(InvoiceData, id=record_id)
            serializer = InvoiceDataSerializer(invoice)
            response = requests.get(serializer.data["attatched_invoice"])
            if response.status_code == 200:
                file_content = response.content
                content_type = response.headers.get("Content-Type", f"application/pdf")
                file_response = HttpResponse(file_content, content_type=content_type)
                file_response["Content-Disposition"] = (
                    f'attachment; filename={f"{invoice.invoice_number}_invoice.pdf"}'
                )
                return file_response
            else:
                return HttpResponse(
                    "Failed to download the file", status=response.status_code
                )

        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to download invoice."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "superadmin", "finance")])
def add_vendor(request):
    data = request.data
    email = data.get("email", "").strip().lower()
    vendor_id = data.get("vendor", "")
    phone = data.get("phone", "")

    try:
        response = create_or_update_vendor_user(email, vendor_id, phone)
        return Response(response, status=status.HTTP_200_OK)
    except ValueError as e:
        return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response(
            {"detail": "An unexpected error occurred."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance", "superadmin")])
def get_all_vendors(request):
    try:
        vendors = Vendor.objects.all().order_by("-created_at")
        serializer = VendorSerializer(vendors, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    except Exception as e:
        print(str(e))
        return Response(
            {"detail": f"Error fetching vendors: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("finance", "ctt_pmo", "pmo")])
def get_zoho_vendors(request):
    try:
        vendors = ZohoVendor.objects.all()
        serializer = ZohoVendorSerializer(vendors, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    except Exception as e:
        print(str(e))
        return Response(
            {"detail": f"Error fetching vendors: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes(
    [IsAuthenticated, IsInRoles("pmo", "vendor", "superadmin", "finance", "sales")]
)
def get_all_purchase_orders(request):
    try:
        all_purchase_orders = PurchaseOrderGetSerializer(
            PurchaseOrder.objects.filter(
                Q(created_time__year__gte=2024)
                | Q(purchaseorder_number__in=purchase_orders_allowed)
            ),
            many=True,
        ).data

        return Response(all_purchase_orders, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance", "sales")])
def get_all_purchase_orders_for_pmo(request):
    try:
        all_purchase_orders = PurchaseOrderGetSerializer(
            PurchaseOrder.objects.all(),
            many=True,
        ).data

        pmos_allowed = json.loads(env("PMOS_ALLOWED_TO_VIEW_ALL_INVOICES_AND_POS"))
        if not request.user.username in pmos_allowed:
            all_purchase_orders = [
                purchase_order
                for purchase_order in all_purchase_orders
                if purchase_order["cf_invoice_approver_s_email"].strip().lower()
                == request.user.username.strip().lower()
            ]

        return Response(all_purchase_orders, status=status.HTTP_200_OK)

    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance", "sales", "leader")])
def get_all_invoices(request):
    try:
        all_invoices = fetch_invoices_db()
        project_id = request.query_params.get("project_id")
        project_type = request.query_params.get("projectType")
        if project_id and project_type:
            purchase_order_ids = get_purchase_order_ids_for_project(
                project_id, project_type
            )
            all_invoices = [
                invoice
                for invoice in all_invoices
                if invoice["purchase_order_id"] in purchase_order_ids
            ]
        return Response(all_invoices, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance", "sales")])
def get_invoice(request, invoice_id):
    try:
        # Get the invoice with related bill in a single query
        invoice = InvoiceData.objects.select_related("bill").get(id=invoice_id)
        # Use the updated serializer that includes the bill relationship
        invoice_serializer = InvoiceDataSerializer(invoice)
        # Return the serialized data which already includes the bill information
        return Response(invoice_serializer.data)
    except Exception as e:
        print(str(e))
        return Response({"error": ""}, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance", "sales")])
def get_invoices_for_pmo(request):
    try:
        all_invoices = fetch_invoices_db()
        pmos_allowed = json.loads(env("PMOS_ALLOWED_TO_VIEW_ALL_INVOICES_AND_POS"))
        if not request.user.username in pmos_allowed:
            all_invoices = [
                invoice
                for invoice in all_invoices
                if invoice["approver_email"].strip().lower() == request.user.username
            ]
        return Response(all_invoices, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance", "sales")])
def get_pending_invoices_for_pmo(request):
    try:
        all_invoices = fetch_invoices_db()
        pmos_allowed = json.loads(env("PMOS_ALLOWED_TO_VIEW_ALL_INVOICES_AND_POS"))
        if not request.user.username in pmos_allowed:
            all_invoices = [
                invoice
                for invoice in all_invoices
                if invoice["approver_email"].strip().lower() == request.user.username
                and invoice["status"] == "pending"
                and invoice["bill"] is None
            ]
        return Response(all_invoices, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance", "sales")])
def get_invoices_for_sales(request):
    try:
        all_invoices = fetch_invoices_db()
        return Response(all_invoices, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance")])
def get_invoices_by_status(request, status):
    try:
        all_invoices = fetch_invoices_db()
        res = []
        for invoice_data in all_invoices:
            if status == "approved":
                if invoice_data["bill"]:
                    if (
                        "status" in invoice_data["bill"]
                        and not invoice_data["bill"]["status"] == "paid"
                    ):
                        res.append(invoice_data)
                elif invoice_data["status"] == "approved":
                    res.append(invoice_data)
            elif status == "paid":
                if invoice_data["bill"] and invoice_data["bill"]["status"] == "paid":
                    res.append(invoice_data)

        return Response(res, status=200)

    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to load"}, status=400)


def get_purchase_order_ids_for_project(project_id, project_type):
    purchase_orders = []
    if project_type == "skill_training" or project_type == "SEEQ":
        purchase_orders = PurchaseOrder.objects.filter(
            schedular_project__id=project_id
        ).values_list("purchaseorder_id", flat=True)
    elif project_type == "CAAS" or project_type == "COD":
        purchase_orders = PurchaseOrder.objects.filter(
            caas_project__id=project_id
        ).values_list("purchaseorder_id", flat=True)
    purchase_order_ids = list(purchase_orders)
    return purchase_order_ids


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance", "sales")])
def get_invoices_by_status_for_founders(request, status):
    try:
        all_invoices = fetch_invoices_db()
        project_id = request.query_params.get("project_id")
        project_type = request.query_params.get("projectType")
        if project_id and project_type:
            purchase_order_ids = get_purchase_order_ids_for_project(
                project_id, project_type
            )
        res = []
        status_counts = defaultdict(int)

        for invoice_data in all_invoices:
            if (
                project_id
                and invoice_data["purchase_order_id"] not in purchase_order_ids
            ):
                continue
            # if status == "in_review":
            if not invoice_data["bill"] and invoice_data["status"] == "in_review":
                status_counts["in_review"] += 1
                if status == "in_review" or status == "all":
                    res.append(invoice_data)
            # elif status == "approved":
            if not invoice_data["bill"] and invoice_data["status"] == "approved":
                status_counts["approved"] += 1
                if status == "approved" or status == "all":
                    res.append(invoice_data)
            # elif status == "rejected":
            if not invoice_data["bill"] and invoice_data["status"] == "rejected":
                status_counts["rejected"] += 1
                if status == "rejected" or status == "all":
                    res.append(invoice_data)
            # if status == "accepted":
            if invoice_data["bill"]:
                if (
                    "status" in invoice_data["bill"]
                    and not invoice_data["bill"]["status"] == "paid"
                ):
                    status_counts["accepted"] += 1
                    if status == "accepted" or status == "all":
                        res.append(invoice_data)
            # elif status == "paid":
            if (
                invoice_data["bill"]
                and invoice_data["bill"]["status"] == "paid"
                or status == "all"
            ):
                status_counts["paid"] += 1
                if status == "paid" or status == "all":
                    res.append(invoice_data)
        return Response({"invoice_counts": status_counts, "invoices": res}, status=200)

    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to load"}, status=400)


@api_view(["PUT"])
@permission_classes(
    [IsAuthenticated, IsInRoles("pmo", "vendor", "superadmin", "finance")]
)
def edit_vendor_existing(request, vendor_id):
    try:
        vendor = Vendor.objects.get(id=vendor_id)
        data = request.data
        email = data.get("email", "").strip().lower()
        vendor_id = data.get("vendor", "")
        phone = data.get("phone", "")
        existing_user = (
            User.objects.filter(username=email).exclude(username=vendor.email).first()
        )
        if existing_user:
            return Response(
                {"error": "User with this email already exists."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        vendor_details = get_vendor(vendor_id)
        name = vendor_details["contact_name"]
        vendor.user.user.username = email
        vendor.user.user.email = email
        vendor.user.user.save()
        vendor.email = email
        vendor.name = name
        vendor.phone = phone
        vendor.vendor_id = vendor_id

        vendor.save()

        return Response(
            {"message": "Vendor updated successfully!"}, status=status.HTTP_200_OK
        )
    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to update vendor"}, status=status.HTTP_404_NOT_FOUND
        )


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "superadmin", "finance")])
def update_invoice_allowed(request, vendor_id):
    try:
        vendor = Vendor.objects.get(id=vendor_id)
    except Vendor.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    serializer = VendorEditSerializer(vendor, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance")])
def update_invoice_status(request, invoice_id):
    try:
        try:
            invoice = InvoiceData.objects.get(pk=invoice_id)
        except InvoiceData.DoesNotExist:
            return Response({"error": "Invoice does not exist"}, status=404)
        new_status = request.data.get("status")
        if new_status not in ["rejected", "approved"]:
            return Response({"error": "Invalid status"}, status=400)
        vendor = Vendor.objects.get(vendor_id=invoice.vendor_id)
        invoice.status = new_status
        invoice.save()
        comment = request.data.get("comment", "")
        approval = InvoiceStatusUpdate.objects.create(
            invoice=invoice,
            status=new_status,
            comment=comment,
            user=request.user,
        )
        approval.save()
        invoice_data = get_invoice_data_for_pdf(
            InvoiceDataSerializer(invoice).data, vendor.hsn_or_sac
        )
        if new_status == "approved":
            email_body_message = render_to_string(
                "vendors/approve_invoice.html",
                {
                    **invoice_data,
                    "comment": comment,
                    "approved_by": request.user.username,
                    "entity_info": {
                        f"This payment will be processed through {vendor.zoho_vendor.entity.name} ({vendor.zoho_vendor.entity.country} entity)"
                    },
                },
            )

            send_mail_templates_with_attachment(
                "invoice_pdf.html",
                json.loads(env("FINANCE_EMAIL")),
                f"{ vendor.zoho_vendor.entity.country} Meeraq | Invoice Approved - {invoice_data['purchase_order_no']} -> {invoice_data['vendor_name']} ",
                {"invoice": invoice_data},
                email_body_message,
                [env("BCC_EMAIL")],
                vendor.is_upload_invoice_allowed,
            )
        else:
            send_mail_to = (
                invoice.vendor_email
                if env("ENVIRONMENT") == "PRODUCTION"
                else "tech@meeraq.com"
            )
            send_mail_templates(
                "vendors/reject_invoice.html",
                [send_mail_to],
                "Meeraq - Invoice Rejected",
                {"vendor_name": invoice.vendor_name, "comment": comment},
                [],
            )
        return Response({"message": f"Invoice {invoice.invoice_number} {new_status}."})
    except Exception as e:
        print(str(e))
        return Response({"error": "Something went wrong"}, status=500)


@api_view(["GET"])
@permission_classes(
    [IsAuthenticated, IsInRoles("pmo", "finance", "vendor", "superadmin")]
)
def get_invoice_updates(request, invoice_id):
    try:
        updates = InvoiceStatusUpdate.objects.filter(invoice_id=invoice_id).order_by(
            "-created_at"
        )
        serializer = InvoiceStatusUpdateGetSerializer(updates, many=True)
        return Response(serializer.data)
    except InvoiceStatusUpdate.DoesNotExist:
        return Response(status=404)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance", "superadmin")])
def get_vendor_details_from_zoho(request, vendor_id):
    try:

        vendor = Vendor.objects.get(vendor_id=vendor_id)

        user = vendor.user.user
        user_data = get_user_data(user)
        if user_data:

            res = {
                "vendor": user_data,
                "zoho_vendor": ZohoVendorSerializer(
                    ZohoVendor.objects.get(contact_id=vendor_id)
                ).data,
            }
            return Response(res)
    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to get data."}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance")])
def get_po_number_to_create(request):
    try:
        entity = request.query_params.get("entity")
        project_type = request.query_params.get("project_type")
        purchase_orders = PurchaseOrderGetSerializer(
            PurchaseOrder.objects.all(),
            many=True,
        ).data
        production = True if env("ENVIRONMENT") == "PRODUCTION" else False
        current_financial_year = get_current_financial_year()

        entity = Entity.objects.filter(id=entity).first()

        regex_to_match = f"BSC/PO/{entity.suffix}/{current_financial_year}/{ '' if production else 'Testing/'}"
        new_po_number = generate_new_po_number(
            purchase_orders, regex_to_match, production
        )

        return Response({"new_po_number": new_po_number})
    except Exception as e:
        print(str(e))
        return Response(status=400)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance", "sales")])
def get_client_invoice_number_to_create(request):
    try:
        start_date, end_date = get_current_month_start_and_end_date()
        query_params = f"&created_date_start={start_date}&created_date_end={end_date}"
        invoices = fetch_client_invoices(organization_id, query_params)
        new_invoice_number = generate_new_invoice_number(invoices)
        return Response({"new_invoice_number": new_invoice_number})
    except Exception as e:
        print(str(e))
        return Response(status=400)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance", "sales")])
def get_so_number_to_create(request):
    try:
        sales_orders = SalesOrderGetSerializer(SalesOrder.objects.all(), many=True).data
        current_financial_year = get_current_financial_year()
        regex_to_match = None
        production = True if env("ENVIRONMENT") == "PRODUCTION" else False

        # Get entity information
        entity_id = request.query_params.get("entity")
        entity = Entity.objects.filter(id=entity_id).first()

        suffix = entity.suffix
        regex_to_match = (
            f"BSC/{suffix}/{current_financial_year}/{'' if production else 'Testing/'}"
        )

        # Generate the new SO number
        new_so_number = generate_new_so_number(
            sales_orders, regex_to_match, production, entity
        )
        return Response({"new_so_number": new_so_number})
    except Exception as e:
        print(str(e))
        return Response(status=400)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance", "sales")])
def get_credit_note_number_to_create(request, brand):
    try:
        # Get current financial year
        current_financial_year = get_current_financial_year()

        # Format prefix based on brand
        brand_prefix = "CTT" if brand.lower() == "ctt" else "MRQ"
        prefix = f"CRN/{brand_prefix}/{current_financial_year}/"

        # Find the last credit note number for this brand and financial year
        credit_notes = CreditNote.objects.filter(
            brand=brand.lower(), credit_note_number__startswith=prefix
        )

        if not credit_notes.exists():
            # If no credit notes exist for this brand and financial year
            new_number = f"{prefix}001"
        else:
            # Get the last credit note number
            last_credit_note = credit_notes.aggregate(Max("credit_note_number"))[
                "credit_note_number__max"
            ]

            try:
                # Extract the sequence number from the last credit note
                last_sequence = int(last_credit_note.split("/")[-1])
                # Generate new sequence number with leading zeros
                new_sequence = str(last_sequence + 1).zfill(3)
                new_number = f"{prefix}{new_sequence}"
            except (ValueError, IndexError, AttributeError):
                # If there's any error parsing the last number, start from 001
                new_number = f"{prefix}001"

        return Response(
            {"new_cn_number": new_number, "financial_year": current_financial_year}
        )

    except Exception as e:
        print(f"Error generating credit note number: {str(e)}")
        return Response({"error": "Failed to generate credit note number"}, status=400)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance")])
def update_purchase_order_status(request, purchase_order_id, status):
    try:
        purchase_order_data = PurchaseOrder.objects.filter(
            purchaseorder_id=purchase_order_id
        ).first()
        if not purchase_order_data:
            return None, status.HTTP_404_NOT_FOUND, "Purchase order not found"

        purchase_order_data.status = status
        purchase_order_data.order_status = status
        purchase_order_data.current_sub_status = status
        purchase_order_data.save()
        return Response({"message": f"Purchase Order changed to {status}."})
    except Exception as e:
        print(str(e))
        return Response(status=404)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_purchase_order_for_outside_vendors(request):
    try:
        entity = None

        entity = Entity.objects.get(id=request.data.get("entity"))

        JSONString = json.loads(request.data.get("JSONString"))

        success = create_singapore_purchase_order(request, JSONString, entity)
        if success:
            return Response({"message": "Purchase Order created successfully."})

        else:
            return Response({"error": "Failed to create purchase order."}, status=500)

    except Exception as e:
        print(str(e))
        return Response(status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_the_invoices_counts(request):
    try:
        all_invoices = fetch_invoices_db()
        res = []
        status_counts = defaultdict(int)
        for invoice_data in all_invoices:
            if not invoice_data["bill"] and invoice_data["status"] == "in_review":
                status_counts["in_review"] += 1
            if not invoice_data["bill"] and invoice_data["status"] == "approved":
                status_counts["approved"] += 1
            if not invoice_data["bill"] and invoice_data["status"] == "rejected":
                status_counts["rejected"] += 1
            if invoice_data["bill"]:
                if (
                    "status" in invoice_data["bill"]
                    and not invoice_data["bill"]["status"] == "paid"
                ):
                    status_counts["accepted"] += 1
            if invoice_data["bill"] and invoice_data["bill"]["status"] == "paid":
                status_counts["paid"] += 1
            res.append(invoice_data)
        return Response({"invoice_counts": status_counts, "invoices": res}, status=200)

    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to load"}, status=400)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_individual_vendor_data(request, vendor_id):
    data, http_status = fetch_and_filter_purchase_orders(vendor_id)
    return Response(data, status=http_status)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_invoices_for_vendor(request, vendor_id, purchase_order_id):
    invoice_res, http_status = fetch_and_process_invoices(vendor_id, purchase_order_id)
    return Response(invoice_res, status=http_status)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_sales_orders(request):
    try:
        search_text = request.query_params.get("search_text", "")
        query_params = (
            f"&salesorder_number_contains={search_text}" if search_text else ""
        )
        all_sales_orders = SalesOrderGetSerializer(
            SalesOrder.objects.all(), many=True
        ).data
        # res = get_sales_orders_with_project_details(all_sales_orders)
        return Response(all_sales_orders, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_sales_persons_sales_orders(request, sales_person_id):
    try:
        all_sales_orders = SalesOrderGetSerializer(
            SalesOrder.objects.filter(salesperson_id=sales_person_id),
            many=True,
        ).data
        return Response(all_sales_orders, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_sales_orders_of_project(request, project_id, project_type):
    try:
        all_sales_orders = get_so_for_project(project_id, project_type)
        return Response(all_sales_orders)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_client_invoices_of_project(request, project_id, project_type):
    try:
        all_sales_orders = get_so_for_project(project_id, project_type)
        so_ids = set()
        for sales_order in all_sales_orders:
            so_ids.add(sales_order["salesorder_id"])

        client_invoices = ClientInvoice.objects.filter(
            sales_order__salesorder_id__in=list(so_ids)
        )
        serializer = ClientInvoiceSerializer(client_invoices, many=True)

        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_ctt_sales_orders(request):
    try:
        participant_email = request.query_params.get("participant_email")
        batch_name = request.query_params.get("batch_name")
        all_sales_orders = []
        if participant_email and batch_name:
            all_sales_orders = SalesOrderGetSerializer(
                SalesOrder.objects.filter(
                    Q(salesorder_number__icontains="CTT")
                    | Q(salesorder_number__icontains="ctt")
                    | Q(salesorder_number__icontains="Ctt"),
                    Q(custom_field_hash__cf_ctt_batch__contains=batch_name),
                    Q(zoho_customer__email=participant_email),
                ),
                many=True,
            ).data
        else:

            all_sales_orders = SalesOrderGetSerializer(
                SalesOrder.objects.filter(
                    Q(salesorder_number__icontains="CTT")
                    | Q(salesorder_number__icontains="ctt")
                    | Q(salesorder_number__icontains="Ctt")
                ),
                many=True,
            ).data
        return Response(all_sales_orders)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_ctt_client_invoices_for_participant(request, participant_email, batch_name):
    try:
        client_invoices = ClientInvoice.objects.filter(
            custom_field_hash__cf_ctt_batch__contains=batch_name,
            zoho_customer__email=participant_email,
            sales_order__custom_field_hash__cf_ctt_batch__contains=batch_name,
        )

        all_client_invoices = ClientInvoiceGetSerializer(
            client_invoices, many=True
        ).data
        return Response(all_client_invoices)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_ctt_client_invoices(request):
    try:
        salesorder_id = request.query_params.get("salesorder_id")
        all_client_invoices = []
        if salesorder_id:
            all_client_invoices = ClientInvoiceGetSerializer(
                ClientInvoice.objects.filter(
                    Q(sales_order__salesorder_id=salesorder_id)
                    | Q(sales_order__salesorder_id=salesorder_id)
                    | Q(sales_order__salesorder_id=salesorder_id)
                ),
                many=True,
            ).data
        else:
            all_client_invoices = ClientInvoiceGetSerializer(
                ClientInvoice.objects.filter(
                    Q(sales_order__salesorder_number__icontains="CTT")
                    | Q(sales_order__salesorder_number__icontains="ctt")
                    | Q(sales_order__salesorder_number__icontains="Ctt")
                ),
                many=True,
            ).data
        return Response(all_client_invoices)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_ctt_purchase_orders_lxp(request):
    try:
        # Filter purchase orders where the order starts with 'ctt'
        purchase_orders = PurchaseOrder.objects.filter(
            purchaseorder_number__startswith="CTT"
        ).distinct()

        # Serialize purchase orders
        serializer = PurchaseOrderSerializer(purchase_orders, many=True)

        return Response(serializer.data)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_ctt_invoices(request):
    try:
        # Fetch faculty emails from the 'ctt' database
        invoices = InvoiceData.objects.filter(
            Q(created_at__year__gte=2024)
            | Q(purchase_order_no__in=purchase_orders_allowed),
            purchase_order_no__startswith="CTT",
        )
        invoice_serializer = InvoiceDataGetSerializer(invoices, many=True)
        return Response(invoice_serializer.data)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def get_sales_order_line_items(line_items):
    """Format line items for the sales order template"""
    formatted_items = []
    for item in line_items:
        formatted_item = {
            "description": item.description,
            "due_date": (
                item.custom_field_hash.get("cf_due_date")
                if isinstance(item.custom_field_hash, dict)
                else None
            ),
            "quantity": f"{float(item.quantity):.2f}",
            "rate": f"{float(item.rate):.2f}",
            "amount": f"{float(item.quantity * item.rate):.2f}",
        }
        formatted_items.append(formatted_item)
    return formatted_items


class DownloadSalesOrder(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo", "vendor", "finance")]

    def get(self, request, id):
        try:
            # Get the sales order
            sales_order = get_object_or_404(SalesOrder, id=id)

            # Calculate sub total from line items
            # Convert total amount to words
            amount_to_words = amount_convert_to_words(
                Decimal(sales_order.total), sales_order.currency_code
            )

            logo = "templates/logos/invoice_logo.png"
            image_base64 = None
            if logo:
                with open(logo, "rb") as image_file:
                    image_base64 = base64.b64encode(image_file.read()).decode("utf-8")

            # Format total with 2 decimal places
            formatted_total = f"{float(sales_order.total):.2f}"

            # Prepare data for the template
            data = {
                # Company Information
                "image_base64": image_base64,
                "organization_name": sales_order.entity.name,
                "organization_billing_address": sales_order.entity.billing_address,
                # Sales Order Details
                "sales_order_number": sales_order.salesorder_number,
                "order_date": (
                    sales_order.date.strftime("%d/%m/%Y") if sales_order.date else ""
                ),
                "ref_number": sales_order.reference_number,
                "sales_person": (
                    sales_order.salesperson_name if sales_order.salesperson_name else ""
                ),
                # Customer Information
                "customer_name": sales_order.zoho_customer.contact_name,
                "customer_address": (
                    format_address(sales_order.zoho_customer.billing_address)
                    if sales_order.zoho_customer
                    else "N/A"
                ),
                # Place of supply
                "place_of_supply": sales_order.place_of_supply,
                # Line Items
                "line_items": get_sales_order_line_items(
                    sales_order.so_line_items.all()
                ),
                # Totals
                "total_amount": f"{sales_order.currency_symbol} {formatted_total}",
                "amount_to_words": amount_to_words,
                # Additional Information
                "currency_code": sales_order.currency_code,
                "currency_symbol": sales_order.currency_symbol,
            }

            # Generate PDF
            email_message = render_to_string(
                "finance/sales_order.html",  # Your sales order template
                data,
            )

            pdf = pdfkit.from_string(email_message, False, configuration=pdfkit_config)

            # Prepare response
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = (
                f"attachment; filename={sales_order.salesorder_number}_sales_order.pdf"
            )

            return response

        except Exception as e:
            print(f"Error generating sales order PDF: {str(e)}")
            return Response(
                {"error": "Failed to download sales order."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


def get_zoho_sales_order(salesorder_id, access_token):
    """
    Fetch sales order data from Zoho Books API.

    Args:
        salesorder_id: Zoho Sales Order ID
        access_token: Zoho API access token

    Returns:
        dict: API response data or None if failed
    """
    try:
        api_url = (
            f"{base_url}/salesorders/{salesorder_id}?organization_id={organization_id}"
        )
        auth_header = {"Authorization": f"Bearer {access_token}"}
        response = requests.get(api_url, headers=auth_header)

        if response.status_code == 200:
            return response.json().get("salesorder")
        return None
    except Exception as e:
        print(f"Error fetching Zoho sales order: {str(e)}")
        return None


def get_existing_sales_order_data(sales_order):
    """
    Get additional data from existing sales order.

    Args:
        sales_order: SalesOrder instance

    Returns:
        dict: Additional sales order data
    """
    return {
        "gm_sheet": sales_order.gm_sheet.id if sales_order.gm_sheet else None,
        "background": sales_order.background,
        "designation": sales_order.designation,
        "linkedin_profile": sales_order.linkedin_profile,
        "referred_by": sales_order.referred_by,
        "companies_worked_in": sales_order.companies_worked_in,
        "performance_evaluation": sales_order.performance_evaluation,
        "entity": sales_order.entity.id if sales_order.entity else None,
        "project_name": sales_order.project_name,
    }


def get_sales_order_data_util(salesorder_id):
    try:
        sales_order = SalesOrder.objects.filter(salesorder_id=salesorder_id).first()
        if not sales_order:
            return None, status.HTTP_404_NOT_FOUND, "Sales order not found"
        # For non-India entity
        serializer = SalesOrderSerializer(sales_order)
        line_items = V2SalesOrderLineItemSerializer(
            SalesOrderLineItem.objects.filter(salesorder=sales_order), many=True
        ).data
        paid_so_line_item_ids = get_paid_so_item_ids(sales_order)

        complete_data = {
            **serializer.data,
            "paid_so_line_item_ids": paid_so_line_item_ids,
            "line_items": line_items,
        }

        return complete_data, status.HTTP_200_OK, None

    except Exception as e:
        print(f"Error in get_sales_order_data_util: {str(e)}")
        return None, status.HTTP_500_INTERNAL_SERVER_ERROR, str(e)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_sales_order_data(request, salesorder_id):
    try:
        data, status_code, error = get_sales_order_data_util(salesorder_id)
        if error:
            return Response({"error": error}, status=status_code)
        return Response(data, status=status_code)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_sales_order_data_from_purchase_order_id(request, purchaseorder_id):
    try:
        purchase_order = PurchaseOrder.objects.get(purchaseorder_id=purchaseorder_id)
        # Get sales order IDs
        sales_order_ids = []
        if purchase_order.caas_project:
            sales_order_ids.extend(
                purchase_order.caas_project.salesorder_set.values_list(
                    "salesorder_id", flat=True
                )
            )
        if purchase_order.schedular_project:
            sales_order_ids.extend(
                purchase_order.schedular_project.salesorder_set.values_list(
                    "salesorder_id", flat=True
                )
            )
        if not sales_order_ids:
            return Response(
                {"error": "Sales order not found"}, status=status.HTTP_404_NOT_FOUND
            )
        salesorder_id = sales_order_ids[0]
        sales_order = SalesOrder.objects.filter(salesorder_id=salesorder_id).first()
        if not sales_order:
            return Response(
                {"error": "Sales order not found"}, status=status.HTTP_404_NOT_FOUND
            )

        # Get sales order data
        data, status_code, error = get_sales_order_data_util(salesorder_id)
        data["po_line_items_ope"] = PurchaseOrderLineItemSerializer(
            purchase_order.po_line_items, many=True
        ).data
        if error:
            return Response({"error": error}, status=status_code)
        return Response(data, status=status_code)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_customers(request):
    try:
        entity_id = request.query_params.get("entity")
        entity = Entity.objects.filter(id=entity_id).first()
        customers = ZohoCustomer.objects.filter(entity=entity)
        customers = ZohoCustomerSerializer(customers, many=True).data
        return Response(customers, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_customer_details(request, customer_id):
    try:
        generate_so_without_invoice = request.query_params.get(
            "generate_so_without_invoice"
        )

        zoho_vendor = None

        if generate_so_without_invoice:
            zoho_vendor = ZohoCustomer.objects.filter(contact_id=customer_id).first()
        else:
            zoho_vendor = ZohoCustomer.objects.filter(id=customer_id).first()
        if zoho_vendor:
            zoho_vendor = ZohoCustomerSerializer(zoho_vendor).data
        else:
            return Response({"error": "Failed to get data."}, status=500)
        return Response(zoho_vendor)
    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to get data."}, status=500)


def get_totals_for_client_invoice_line_items(line_items):
    total = 0
    sub_total = 0
    sub_total_inclusive_of_tax = 0
    tax_total = 0
    # total_quantity = 0

    for line_item in line_items:
        # Accessing each attribute directly as it's a model instance
        item_total = float(getattr(line_item, "item_total", 0))
        tax_percentage = float(getattr(line_item, "tax_percentage", 0))
        item_total_inclusive_of_tax = float(
            getattr(line_item, "item_total_inclusive_of_tax", 0)
        )
        # item_quantity = float(getattr(line_item, "quantity", 0))

        # Calculating tax for each item based on total and sub_total
        item_tax = item_total * (tax_percentage / 100)
        item_sub_total = item_total - item_tax

        # Aggregate totals
        total += item_total
        sub_total += item_sub_total
        sub_total_inclusive_of_tax += item_total_inclusive_of_tax
        tax_total += item_tax
        # total_quantity += item_quantity

    return {
        "total": total,
        "sub_total": sub_total,
        "sub_total_inclusive_of_tax": sub_total_inclusive_of_tax,
        "tax_total": tax_total,
        # "total_quantity": total_quantity,
    }


def get_quantity_invoices(line_items):
    quantiy_invocied_in_items = {}
    for line_item in line_items:
        item_data = Decimal(getattr(line_item, "quantity", 0))
        quantiy_invocied_in_items[line_item.salesorder_item_id] = item_data

    return quantiy_invocied_in_items


def get_po_quantity_invoices(line_items):
    quantiy_invocied_in_items = {}
    for line_item in line_items:
        item_data = Decimal(getattr(line_item, "quantity", 0))
        quantiy_invocied_in_items[line_item.purchaseorder_item_id] = item_data

    return quantiy_invocied_in_items


def get_sales_order_totals(line_items):
    total = 0
    sub_total = 0
    sub_total_inclusive_of_tax = 0
    tax_total = 0
    total_quantity = 0
    total_invoiced_quantity = 0

    for line_item in line_items:
        # Accessing each attribute directly as it's a model instance
        item_total = float(getattr(line_item, "item_total", 0))
        item_sub_total = float(getattr(line_item, "item_sub_total", 0))
        item_total_inclusive_of_tax = float(
            getattr(line_item, "item_total_inclusive_of_tax", 0)
        )
        item_quantity = float(getattr(line_item, "quantity", 0))
        quantity_invoiced = float(getattr(line_item, "quantity_invoiced", 0))

        # Calculating tax for each item based on total and sub_total
        item_tax = item_total - item_sub_total

        # Aggregate totals
        total += item_total
        sub_total += item_sub_total
        sub_total_inclusive_of_tax += item_total_inclusive_of_tax
        tax_total += item_tax
        total_quantity += item_quantity
        total_invoiced_quantity += quantity_invoiced

    return {
        "total": total,
        "sub_total": sub_total,
        "sub_total_inclusive_of_tax": sub_total_inclusive_of_tax,
        "tax_total": tax_total,
        "total_quantity": total_quantity,
        "total_invoiced_quantity": total_invoiced_quantity,
    }


def get_sales_order_totals_client_invoice(line_items):
    total = 0
    sub_total = 0
    sub_total_inclusive_of_tax = 0
    tax_total = 0
    total_quantity = 0
    total_invoiced_quantity = 0

    for line_item in line_items:
        # Accessing each attribute directly as it's a model instance
        item_total = float(getattr(line_item, "item_total", 0))
        item_sub_total = float(getattr(line_item, "item_sub_total", 0))
        # item_total_inclusive_of_tax = float(
        #     getattr(line_item, "item_total_inclusive_of_tax", 0)
        # )
        item_quantity = float(getattr(line_item, "quantity", 0))
        quantity_invoiced = float(getattr(line_item, "quantity_invoiced", 0))

        # Calculating tax for each item based on total and sub_total
        item_tax = item_total - item_sub_total

        # Aggregate totals
        total += item_total
        sub_total += item_sub_total
        # sub_total_inclusive_of_tax += item_total_inclusive_of_tax
        tax_total += item_tax
        total_quantity += item_quantity
        total_invoiced_quantity += quantity_invoiced

    return {
        "total": total,
        "sub_total": sub_total,
        "sub_total_inclusive_of_tax": sub_total_inclusive_of_tax,
        "tax_total": tax_total,
        "total_quantity": total_quantity,
        "total_invoiced_quantity": total_invoiced_quantity,
    }


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_invoice(request):
    try:
        with transaction.atomic():
            entity = Entity.objects.filter(id=request.data.get("entity")).first()
            purchase_orders = []
            if request.data.get("po_ids", []):
                purchase_orders = PurchaseOrder.objects.filter(
                    id__in=json.loads(request.data.get("po_ids"))
                )

            JSONString = json.loads(request.data.get("JSONString"))

            formatted_custom_fields_data = build_custom_fields_and_hash(
                JSONString.get("custom_fields")
            )
            JSONString["custom_fields"] = formatted_custom_fields_data["custom_fields"]
            JSONString["custom_field_hash"] = formatted_custom_fields_data[
                "custom_field_hash"
            ]

            if request.data.get("without_sales_order", None):
                z_c = ZohoCustomer.objects.filter(
                    contact_id=request.data.get("zoho_customer")
                ).first()
                JSONString["zoho_customer"] = z_c.id

            serializer = ClientInvoiceSerializer(data=JSONString)

            if serializer.is_valid():
                client_invoice = serializer.save()
                client_invoice.invoice_id = f"M-{client_invoice.id}"
                client_invoice.entity = entity
                for purchase_order in purchase_orders:
                    client_invoice.purchase_orders.add(purchase_order)
                client_invoice.save()
                for line_item in JSONString["line_items"]:
                    line_item_serializer = ClientInvoiceLineItemSerializer(
                        data=line_item
                    )
                    if line_item_serializer.is_valid():
                        instance = line_item_serializer.save()
                        instance.line_item_id = f"M-{instance.id}"
                        instance.save()
                        client_invoice.client_invoice_line_items.add(instance)
                    else:
                        print(
                            JSONString["invoice_number"],
                            line_item_serializer.errors,
                        )

                rate = get_exchange_rate(JSONString.get("currency_code"), "INR")
                line_item_totals = get_totals_for_client_invoice_line_items(
                    client_invoice.client_invoice_line_items.all()
                )
                client_invoice.exchange_rate = rate
                client_invoice.price_precision = 2
                client_invoice.total = line_item_totals["total"]
                client_invoice.sub_total = line_item_totals["sub_total"]
                client_invoice.sub_total_inclusive_of_tax = line_item_totals[
                    "sub_total_inclusive_of_tax"
                ]
                client_invoice.tax_total = line_item_totals["tax_total"]
                # client_invoice.total_quantity = line_item_totals['total_quantity']
                client_invoice.created_time = datetime.now()
                client_invoice.created_date = datetime.now().date()
                client_invoice.last_modified_time = datetime.now()
                sales_order = SalesOrder.objects.filter(
                    salesorder_id=JSONString.get("salesorder_id")
                ).first()
                quantiy_invocied_in_items = get_quantity_invoices(
                    client_invoice.client_invoice_line_items.all()
                )
                # updating so quantity input and if so line item is invoiced
                for line_item in sales_order.so_line_items.all():
                    if line_item.line_item_id in quantiy_invocied_in_items:
                        line_item.is_invoiced = True
                        line_item.quantity_invoiced += quantiy_invocied_in_items[
                            line_item.line_item_id
                        ]
                        line_item.save()
                sales_order.save()
                client_invoice.sales_order = sales_order
                client_invoice.save()
                client_invoice = update_invoice_status_and_balance(client_invoice)
                # updating sales order fields on invoicing
                sales_order.invoices = ClientInvoiceGetSerializer(
                    ClientInvoice.objects.filter(sales_order=sales_order), many=True
                ).data
                sales_order_line_item_totals = get_sales_order_totals_client_invoice(
                    sales_order.so_line_items.all()
                )
                if (
                    sales_order_line_item_totals["total_quantity"]
                    == sales_order_line_item_totals["total_invoiced_quantity"]
                ):
                    sales_order.status = "invoiced"
                    sales_order.invoiced_status = "invoiced"
                else:
                    sales_order.status = "partially_invoiced"
                    sales_order.invoiced_status = "partially_invoiced"
                sales_order.save()
                res_data = ClientInvoiceSerializer(client_invoice).data
                return Response(res_data, status=201)
            else:
                print(serializer.errors)
                return Response({"error": "Invalid data"}, status=404)
    except Exception as e:
        print(str(e))
        return Response(status=404)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def edit_so_invoice(request, invoice_id):
    try:
        print(f"Starting edit_so_invoice for invoice_id: {invoice_id}")
        with transaction.atomic():
            client_invoice_data = ClientInvoice.objects.filter(
                invoice_id=invoice_id
            ).first()
            if not client_invoice_data:
                print("Client invoice not found.")
                return Response({"error": "Client invoice not found."}, status=404)

            entity = Entity.objects.filter(id=request.data.get("entity")).first()
            JSONString = json.loads(request.data.get("JSONString"))
            print(f"Parsed JSONString: {JSONString}")

            if entity.id == int(env("INDIA_ENTITY_ID")):
                access_token = get_access_token(env("ZOHO_REFRESH_TOKEN"))
                if not access_token:
                    print("Access token not found.")
                    return Response(
                        {
                            "error": "Access token not found. Please generate an access token first."
                        },
                        status=400,
                    )

                api_url = f"{base_url}/invoices/{invoice_id}?organization_id={organization_id}"
                auth_header = {"Authorization": f"Bearer {access_token}"}
                response = requests.put(api_url, headers=auth_header, data=request.data)
                print(f"Zoho API response status: {response.status_code}")

                if response.status_code == 200:
                    invoice = response.json().get("invoice")
                    if invoice:
                        print(
                            f"Updating client invoice with ID: {invoice['invoice_id']}"
                        )
                        create_or_update_client_invoice(invoice["invoice_id"])
                    return Response({"message": "Invoice updated successfully."})
                else:
                    print(f"Error from Zoho API: {response.json()}")
                    return Response(
                        {"error": response.json()}, status=response.status_code
                    )
            else:
                formatted_custom_fields_data = build_custom_fields_and_hash(
                    JSONString.get("custom_fields")
                )
                JSONString["custom_fields"] = formatted_custom_fields_data[
                    "custom_fields"
                ]
                JSONString["custom_field_hash"] = formatted_custom_fields_data[
                    "custom_field_hash"
                ]
                print(f"Formatted custom fields: {formatted_custom_fields_data}")

                existing_line_items = (
                    client_invoice_data.client_invoice_line_items.all()
                )
                existing_line_items_dict = {
                    item.line_item_id: item for item in existing_line_items
                }
                existing_line_item_ids = set(existing_line_items_dict.keys())
                print(f"Existing line item IDs: {existing_line_item_ids}")

                new_line_items = JSONString["line_items"]
                incoming_line_item_ids = set(
                    line_item["line_item_id"]
                    for line_item in new_line_items
                    if "line_item_id" in line_item
                )
                print(f"Incoming line item IDs: {incoming_line_item_ids}")

                line_item_ids_to_remove = (
                    existing_line_item_ids - incoming_line_item_ids
                )
                line_item_ids_to_update = (
                    existing_line_item_ids & incoming_line_item_ids
                )
                line_item_ids_to_add = {
                    line_item.get("line_item_id")
                    for line_item in new_line_items
                    if "line_item_id" not in line_item
                    or line_item["line_item_id"] not in existing_line_item_ids
                }
                print(f"Line item IDs to remove: {line_item_ids_to_remove}")
                print(f"Line item IDs to update: {line_item_ids_to_update}")
                print(f"Line item IDs to add: {line_item_ids_to_add}")

                JSONString["line_items"] = process_line_item_custom_fields(
                    JSONString["line_items"]
                )
                print(f"Processed line items: {JSONString['line_items']}")

                quantity_changes = {}

                for line_item in JSONString["line_items"]:
                    line_item_id = line_item.get("line_item_id")
                    salesorder_item_id = line_item.get("salesorder_item_id")
                    if line_item_id in line_item_ids_to_update:
                        existing_line_item = existing_line_items_dict[line_item_id]
                        old_quantity = existing_line_item.quantity
                        line_item_serializer = ClientInvoiceLineItemSerializer(
                            existing_line_item, data=line_item, partial=True
                        )
                        if line_item_serializer.is_valid():
                            updated_line_item = line_item_serializer.save()
                            if salesorder_item_id in quantity_changes:
                                quantity_changes[salesorder_item_id] += (
                                    updated_line_item.quantity - old_quantity
                                )
                            else:
                                quantity_changes[salesorder_item_id] = (
                                    updated_line_item.quantity - old_quantity
                                )
                            print(
                                f"Updated line item ID: {line_item_id}, Quantity change: {quantity_changes[salesorder_item_id]}"
                            )
                        else:
                            print(
                                f"Error in line item serializer: {line_item_serializer.errors}"
                            )
                            return Response(
                                {"error": line_item_serializer.errors}, status=400
                            )
                    elif (
                        line_item_id in line_item_ids_to_add
                        or "line_item_id" not in line_item
                    ):
                        line_item_serializer = ClientInvoiceLineItemSerializer(
                            data=line_item
                        )
                        if line_item_serializer.is_valid():
                            new_line_item = line_item_serializer.save()
                            if not new_line_item.line_item_id:
                                new_line_item.line_item_id = f"M-{new_line_item.id}"
                                new_line_item.save()
                            if salesorder_item_id in quantity_changes:
                                quantity_changes[
                                    salesorder_item_id
                                ] += new_line_item.quantity
                            else:
                                quantity_changes[salesorder_item_id] = (
                                    new_line_item.quantity
                                )
                            client_invoice_data.client_invoice_line_items.add(
                                new_line_item
                            )
                            print(
                                f"Added new line item ID: {new_line_item.line_item_id}, Quantity: {new_line_item.quantity}"
                            )
                        else:
                            print(
                                f"Error in new line item serializer: {line_item_serializer.errors}"
                            )
                            return Response(
                                {"error": line_item_serializer.errors}, status=400
                            )

                for line_item_id in line_item_ids_to_remove:
                    removed_line_item = existing_line_items_dict[line_item_id]
                    if salesorder_item_id in quantity_changes:
                        quantity_changes[
                            removed_line_item.salesorder_item_id
                        ] += -removed_line_item.quantity
                    else:
                        quantity_changes[removed_line_item.salesorder_item_id] = (
                            -removed_line_item.quantity
                        )
                    removed_line_item.delete()
                    print(
                        f"Removed line item ID: {line_item_id}, Quantity change: {quantity_changes[removed_line_item.salesorder_item_id]}"
                    )

                serializer = ClientInvoiceSerializer(
                    client_invoice_data, data=JSONString, partial=True
                )

                if serializer.is_valid():
                    client_invoice = serializer.save()
                    client_invoice.last_modified_time = datetime.now()
                    if (
                        "currency_code" in JSONString
                        and JSONString.get("currency_code")
                        != client_invoice.currency_code
                    ):
                        rate = get_exchange_rate(JSONString.get("currency_code"), "INR")
                        client_invoice.exchange_rate = rate
                        print(f"Exchange rate updated: {rate}")

                    line_item_totals = get_totals_for_client_invoice_line_items(
                        client_invoice.client_invoice_line_items.all()
                    )
                    client_invoice.price_precision = 2
                    client_invoice.total = line_item_totals["total"]
                    client_invoice.sub_total = line_item_totals["sub_total"]
                    client_invoice.sub_total_inclusive_of_tax = line_item_totals[
                        "sub_total_inclusive_of_tax"
                    ]
                    client_invoice.tax_total = line_item_totals["tax_total"]
                    client_invoice.save()
                    client_invoice = update_invoice_status_and_balance(client_invoice)
                    print(f"Client invoice totals updated: {line_item_totals}")

                    sales_order = client_invoice.sales_order
                    sales_order_line_items = sales_order.so_line_items.all()
                    sales_order_line_items_dict = {
                        item.line_item_id: item for item in sales_order_line_items
                    }

                    for line_item_id, quantity_change in quantity_changes.items():
                        so_line_item = sales_order_line_items_dict.get(line_item_id)
                        if so_line_item:
                            so_line_item.quantity_invoiced += quantity_change
                            so_line_item.is_invoiced = (
                                so_line_item.quantity_invoiced >= 0
                            )
                            so_line_item.save()
                            print(
                                f"Sales order line item updated: {line_item_id}, Quantity invoiced: {so_line_item.quantity_invoiced}"
                            )

                    sales_order_line_item_totals = get_sales_order_totals(
                        sales_order.so_line_items.all()
                    )
                    print(
                        f"Sales order line item totals: {sales_order_line_item_totals}"
                    )

                    if (
                        sales_order_line_item_totals["total_quantity"]
                        == sales_order_line_item_totals["total_invoiced_quantity"]
                    ):
                        sales_order.status = "invoiced"
                        sales_order.invoiced_status = "invoiced"
                    elif sales_order_line_item_totals["total_invoiced_quantity"] == 0:
                        sales_order.status = "open"
                        sales_order.invoiced_status = "not_invoiced"
                    else:
                        sales_order.status = "partially_invoiced"
                        sales_order.invoiced_status = "partially_invoiced"

                    sales_order.save()
                    print(
                        f"Sales order status updated: {sales_order.status}, Invoiced status: {sales_order.invoiced_status}"
                    )
                    res_data = ClientInvoiceSerializer(client_invoice).data
                    return Response(res_data, status=200)
                else:
                    print(f"Error in client invoice serializer: {serializer.errors}")
                    return Response({"error": serializer.errors}, status=400)
    except Exception as e:
        print(f"Exception occurred: {str(e)}")
        return Response({"error": str(e)}, status=500)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def edit_sales_order(request, sales_order_id):
    try:
        with transaction.atomic():
            sales_order = None
            entity = Entity.objects.filter(id=request.data.get("entity")).first()
            JSONString = json.loads(request.data.get("JSONString"))
            sales_order_data = SalesOrder.objects.filter(id=sales_order_id).first()
            line_items_data = JSONString.get("line_items")
            formatted_custom_fields_data = build_custom_fields_and_hash(
                JSONString.get("custom_fields")
            )
            JSONString["custom_fields"] = formatted_custom_fields_data["custom_fields"]
            JSONString["custom_field_hash"] = formatted_custom_fields_data[
                "custom_field_hash"
            ]
            # NEW LINE ITEMS
            # EXISTING LINE ITEMS
            # CURRENT REMOVE

            # EXISTING IDS
            # INCOMING LINE ITEM IDS

            # REMOVE = EXISTING IDS - INCOMING LINE ITEM IDS
            # UPDATE = COMMON (EXISTING IDS + INCOMING LINE ITEM IDS)
            # NEW = LINE ITEMS WITH NO LINE ITEM ID

            # Fetch existing line items and their IDs
            existing_line_items = sales_order_data.so_line_items.all()
            existing_line_item_ids = [
                line_item.line_item_id for line_item in existing_line_items
            ]

            # Fetch new line items and their IDs from JSON input
            new_line_items = JSONString["line_items"]
            incoming_line_item_ids = [
                line_item["line_item_id"]
                for line_item in new_line_items
                if "line_item_id" in line_item
            ]

            # Determine line items to remove, update, and add
            line_item_ids_to_remove = list(
                set(existing_line_item_ids) - set(incoming_line_item_ids)
            )
            line_item_ids_to_update = list(
                set(existing_line_item_ids).intersection(incoming_line_item_ids)
            )

            # Process custom fields in JSON line items
            JSONString["line_items"] = process_line_item_custom_fields(
                JSONString["line_items"]
            )
            print("STEP 4")
            all_line_items_data = []

            # Update and Add line items
            for index, line_item in enumerate(JSONString["line_items"]):
                # Update existing line items
                if line_item.get("line_item_id") in line_item_ids_to_update:
                    existing_line_item = existing_line_items.get(
                        line_item_id=line_item["line_item_id"]
                    )
                    line_item_serializer = SalesOrderLineItemSerializer(
                        existing_line_item, data=line_item, partial=True
                    )
                # Add new line items
                elif "line_item_id" not in line_item:
                    line_item_serializer = SalesOrderLineItemSerializer(data=line_item)

                # Save valid line items
                if line_item_serializer.is_valid():
                    instance = line_item_serializer.save()
                    if not instance.line_item_id:
                        instance.line_item_id = f"M-{instance.id}"
                        instance.save()
                    if index == 0:
                        is_booking_amount = json.loads(
                            request.data.get("is_booking_amount", "false")
                        )
                        instance.is_booking_amount = is_booking_amount
                        instance.save()
                    all_line_items_data.append(instance)
                else:
                    print(JSONString["salesorder_number"], line_item_serializer.errors)

            # Remove line items that are no longer needed
            SalesOrderLineItem.objects.filter(
                line_item_id__in=line_item_ids_to_remove
            ).delete()

            # Proceed to validate and save the SalesOrder only if line items were created successfully
            serializer = SalesOrderSerializer(
                sales_order_data, data=JSONString, partial=True
            )
            print("STEP 4")

            # Check if the SalesOrder data is valid
            if serializer.is_valid():
                print("STEP 9")
                # Save the SalesOrder data if valid
                sales_order_data = serializer.save()

                # Retrieve the ZohoCustomer instance if exists
                zoho_customer = ZohoCustomer.objects.filter(
                    id=JSONString.get("zoho_customer")
                ).first()

                sales_user = Sales.objects.filter(
                    sales_person_id=JSONString.get("salesperson_id")
                ).first()

                rate = get_exchange_rate(JSONString.get("currency_code"), "INR")

                # Associate the ZohoCustomer with the SalesOrder
                if zoho_customer:
                    sales_order_data.zoho_customer = zoho_customer
                    sales_order_data.customer_name = zoho_customer.contact_name
                    sales_order_data.contact_category = zoho_customer.contact_category
                    sales_order_data.tax_treatment = zoho_customer.tax_treatment
                sales_order_data.entity = entity
                # Associate line items with the sales order
                for item in all_line_items_data:
                    sales_order_data.so_line_items.add(item)
                sales_order_data.save()
                line_item_totals = get_sales_order_totals(
                    sales_order_data.so_line_items.all()
                )
                sales_order_data.total = line_item_totals["total"]
                sales_order_data.sub_total = line_item_totals["sub_total"]
                sales_order_data.sub_total_inclusive_of_tax = line_item_totals[
                    "sub_total_inclusive_of_tax"
                ]
                sales_order_data.tax_total = line_item_totals["tax_total"]
                sales_order_data.total_quantity = line_item_totals["total_quantity"]
                sales_order_data.current_sub_status = "draft"
                sales_order_data.exchange_rate = rate
                sales_order_data.place_of_supply = (
                    JSONString.get("billing_address").get("state")
                    if JSONString.get("billing_address")
                    else ""
                )
                sales_order_data.salesperson_id = sales_user.sales_person_id
                sales_order_data.salesperson_name = sales_user.name
                sales_order_data.save()

            # Handle invalid SalesOrder data
            else:
                return Response(
                    {
                        "error": "Invalid sales order data",
                        "details": serializer.errors,
                    },
                    status=500,
                )

            if sales_order_data:
                sales_order_data.referred_by = JSONString.get("referredBy", "")
                sales_order_data.linkedin_profile = JSONString.get(
                    "linkedInProfile", ""
                )
                sales_order_data.background = JSONString.get("background", "")
                sales_order_data.designation = JSONString.get("designation", "")
                sales_order_data.companies_worked_in = JSONString.get(
                    "companies_worked_in", ""
                )
                sales_order_data.performance_evaluation = JSONString.get(
                    "performance_evaluation", ""
                )
                sales_order_data.location = JSONString.get("location", "")
                sales_order_data.credential = JSONString.get("credential", "")
                sales_order_data.save()
                gm_sheet_id = request.data.get("gm_sheet", "")
                if gm_sheet_id:
                    try:
                        gm_sheet = GmSheet.objects.get(id=gm_sheet_id)
                        sales_order_data.gm_sheet = gm_sheet
                        sales_order_data.save()
                    except Exception as e:
                        print(str(e))

            return Response({"message": "Sales order updated successfully."})
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=500)


# Metadata Definition
CUSTOM_FIELD_METADATA = {
    "2913550000002949012": {
        "index": 1,
        "label": "Invoicing Type",
        "api_name": "cf_invoicing_type",
        "value_formatted": "",
        "placeholder": "cf_invoicing_type",
        "value": "",
    },
    "2913550000003238005": {
        "index": 2,
        "label": "CTT Batch",
        "api_name": "cf_ctt_batch",
        "value_formatted": "",
        "placeholder": "cf_ctt_batch",
        "value": "",
    },
    "2913550000003277012": {
        "index": 1,
        "label": "CTT Batch",
        "api_name": "cf_ctt_batch",
        "value_formatted": "",
        "placeholder": "cf_ctt_batch",
        "value": "",
    },
}


# Function to retrieve metadata by customfield_id
def get_metadata(customfield_id):
    return CUSTOM_FIELD_METADATA.get(customfield_id, {}).copy()


# Function to process a single field entry and return enhanced data
def process_field(field):
    customfield_id = field["customfield_id"]
    metadata = get_metadata(customfield_id)

    if not metadata:
        # Skip if no metadata found for this ID
        return None

    # Use the initial field's value if provided; fall back to default in metadata
    metadata["value"] = field["value"]
    metadata["value_formatted"] = field["value"]
    metadata["customfield_id"] = customfield_id

    return metadata


# Function to build the final structure
def build_custom_fields_and_hash(initial_data):
    custom_fields = []
    custom_field_hash = {}

    for field in initial_data:
        processed_field = process_field(field)

        if processed_field:
            # Add to custom_fields list
            custom_fields.append(processed_field)

            # Populate the custom_field_hash
            api_name = processed_field["api_name"]
            value = processed_field["value"]
            custom_field_hash[api_name] = value
            custom_field_hash[f"{api_name}_unformatted"] = value

    return {"custom_fields": custom_fields, "custom_field_hash": custom_field_hash}


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_sales_order(request):
    try:
        with transaction.atomic():
            salesorder_created = None
            entity = Entity.objects.filter(id=request.data.get("entity")).first()
            JSONString = json.loads(request.data.get("JSONString"))

            line_items_data = process_line_item_custom_fields(JSONString["line_items"])
            formatted_custom_fields_data = build_custom_fields_and_hash(
                JSONString.get("custom_fields")
            )
            JSONString["custom_fields"] = formatted_custom_fields_data["custom_fields"]
            JSONString["custom_field_hash"] = formatted_custom_fields_data[
                "custom_field_hash"
            ]
            line_items = None
            if line_items_data:
                line_item_serializer = SalesOrderLineItemSerializer(
                    data=line_items_data, many=True
                )

                # Check if line items data is valid
                if line_item_serializer.is_valid():
                    line_items = line_item_serializer.save()  # Save line items
                    for item in line_items:
                        item.line_item_id = f"M-{item.id}"
                        item.save()
                else:
                    # Return error if line items are invalid, and stop further processing
                    return Response(
                        {
                            "error": "Invalid line items data",
                            "details": line_item_serializer.errors,
                        },
                        status=500,
                    )

            # Proceed to validate and save the SalesOrder only if line items were created successfully
            serializer = SalesOrderSerializer(data=JSONString, partial=True)
            # Check if the SalesOrder data is valid
            if serializer.is_valid():
                # Save the SalesOrder data if valid
                sales_order = serializer.save()
                salesorder_created = serializer.data
                sales_order.salesorder_id = f"BSY-{sales_order.id}"
                sales_person = Sales.objects.filter(
                    id=sales_order.salesperson_id
                ).first()
                sales_order.added_by = sales_person.user.user
                sales_order.save()
                # Retrieve the ZohoCustomer instance if exists
                zoho_customer = ZohoCustomer.objects.filter(
                    id=JSONString.get("zoho_customer")
                ).first()
                user_creating_sales_order = Sales.objects.filter(
                    email=request.user.username
                ).first()
                rate = get_exchange_rate(JSONString.get("currency_code"), "INR")

                # Associate the ZohoCustomer with the SalesOrder
                if zoho_customer:
                    sales_order.zoho_customer = zoho_customer
                    sales_order.customer_name = zoho_customer.contact_name
                    sales_order.contact_category = zoho_customer.contact_category
                    sales_order.tax_treatment = zoho_customer.tax_treatment
                sales_order.entity = entity
                sales_order.status = "draft"
                sales_order.order_status = "draft"
                sales_order.invoiced_status = "not_invoiced"
                sales_order.paid_status = "unpaid"
                sales_order.so_line_items.set(line_items)
                line_item_totals = get_sales_order_totals(line_items)
                # Associate line items with the sales order
                sales_order.total = line_item_totals["total"]
                sales_order.sub_total = line_item_totals["sub_total"]
                sales_order.sub_total_inclusive_of_tax = line_item_totals[
                    "sub_total_inclusive_of_tax"
                ]
                sales_order.tax_total = line_item_totals["tax_total"]
                sales_order.total_quantity = line_item_totals["total_quantity"]
                sales_order.current_sub_status = "draft"
                sales_order.created_by_email = user_creating_sales_order.email
                sales_order.created_by_name = user_creating_sales_order.name
                sales_order.created_time = datetime.now()
                sales_order.created_date = datetime.now().date()
                sales_order.last_modified_time = datetime.now()
                sales_order.exchange_rate = rate
                sales_order.place_of_supply = (
                    JSONString.get("billing_address").get("state")
                    if JSONString.get("billing_address")
                    else ""
                )
                sales_order.salesperson_id = sales_person.sales_person_id
                sales_order.salesperson_name = sales_person.name
                sales_order.save()
                salesorder_created = SalesOrderSerializer(sales_order).data
            # Handle invalid SalesOrder data
            else:
                return Response(
                    {
                        "error": "Invalid sales order data",
                        "details": serializer.errors,
                    },
                    status=500,
                )

            gm_sheet_id = request.data.get("gm_sheet", None)
            if gm_sheet_id:
                try:
                    gm_sheet = GmSheet.objects.get(id=gm_sheet_id)
                    sales_order.gm_sheet = gm_sheet
                    sales_order.save()
                    try:
                        gm_sheet.client_name = sales_order.customer_name
                        gm_sheet.save()
                    except Exception as e:
                        print("Failed to update the client name in GmSheet:", str(e))
                except Exception as e:
                    print(str(e))
            so_number = salesorder_created["salesorder_number"]
            customer_name = salesorder_created["customer_name"]
            salesperson_name = salesorder_created["salesperson_name"]
            send_mail_templates(
                "so_emails/sales_order_mail.html",
                (
                    (["finance@coachtotransformation.com"])
                    if env("ENVIRONMENT") == "PRODUCTION"
                    else ["naveen@meeraq.com"]
                ),
                "New Sales Order Created",
                {
                    "so_number": so_number,
                    "customer_name": customer_name,
                    "salesperson": salesperson_name,
                    "total_amount": (
                        round(
                            (
                                Decimal(salesorder_created["total"])
                                - Decimal(salesorder_created["tax_total"])
                                if Decimal(salesorder_created["tax_total"])
                                else Decimal(salesorder_created["total"])
                            ),
                            2,
                        )
                    ),
                    "currency_symbol": salesorder_created["currency_symbol"],
                },
                (
                    (
                        [
                            "rajat@meeraq.com",
                            "sujata@meeraq.com",
                        ]
                    )
                    if env("ENVIRONMENT") == "PRODUCTION"
                    else ["tech@meeraq.com"]
                ),
            )
            so_status = request.data.get("status", "")
            if so_status == "open":

                sales_order.current_sub_status = "open"
                sales_order.status = "open"
                sales_order.order_status = "open"
                sales_order.invoiced_status = "not_invoiced"
                sales_order.paid_status = "unpaid"
                sales_order.save()
            return Response(
                {
                    "message": "SO has been created successfully and Saved as Draft",
                    "salesorder": salesorder_created,
                }
            )

    except Exception as e:
        print(str(e))
        return Response(status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_client_invoices(request):
    try:
        all_client_invoices = ClientInvoiceGetSerializer(
            ClientInvoice.objects.all(), many=True
        ).data
        # fetch_client_invoices(organization_id)
        return Response(all_client_invoices, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_client_invoices_for_project(request):
    try:
        all_client_invoices = ClientInvoiceGetSerializer(
            ClientInvoice.objects.all(), many=True
        ).data
        return Response(all_client_invoices, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_client_invoice_data(request, invoice_id):
    is_id = request.query_params.get("is_id", None)
    if is_id:
        client_invoice_instance = ClientInvoice.objects.filter(id=invoice_id).first()
    else:
        client_invoice_instance = ClientInvoice.objects.filter(
            invoice_id=invoice_id
        ).first()

    serializer = ClientInvoiceSerializer(client_invoice_instance)
    line_items = V2ClientInvoiceLineItemSerializer(
        ClientInvoiceLineItem.objects.filter(clientinvoice=client_invoice_instance),
        many=True,
    ).data
    return Response({**serializer.data, "line_items": line_items}, status=200)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance")])
def update_client_invoice_status(request, invoice_id, status):
    try:
        client_invoice_instance = ClientInvoice.objects.filter(
            invoice_id=invoice_id
        ).first()

        client_invoice_instance.status = "sent"
        client_invoice_instance.current_sub_status = "sent"
        client_invoice_instance.save()
        return Response({"message": f"Client Invoice changed to {status}."})
    except Exception as e:
        print(str(e))
        return Response(status=404)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_sales_order_status(request, sales_order_id, status):
    try:
        sales_order = SalesOrder.objects.get(salesorder_id=sales_order_id)

        if sales_order.status == "draft":
            sales_order.status == "open"
            sales_order.order_status == "open"
            sales_order.save()
        return Response({"message": f"Sales Order changed to {status}."})
    except Exception as e:
        print(str(e))
        return Response(status=404)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def edit_vendor(request, vendor_id):
    try:
        with transaction.atomic():
            # Get vendor and related zoho_vendor
            try:
                vendor = Vendor.objects.get(vendor_id=vendor_id)
                zoho_vendor = vendor.zoho_vendor
            except Vendor.DoesNotExist:
                return Response(
                    {"error": "Vendor not found"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            data = request.data

            # Update ZohoVendor first
            zoho_vendor.contact_name = data.get("name", zoho_vendor.contact_name)
            zoho_vendor.company_name = data.get(
                "company_name", zoho_vendor.company_name
            )
            zoho_vendor.first_name = data.get("first_name", zoho_vendor.first_name)
            zoho_vendor.last_name = data.get("last_name", zoho_vendor.last_name)
            zoho_vendor.email = data.get("email", zoho_vendor.email)
            zoho_vendor.phone = data.get("phone", zoho_vendor.phone)
            zoho_vendor.mobile = data.get("phone", zoho_vendor.mobile)
            zoho_vendor.gst_treatment = data.get(
                "gst_treatment", zoho_vendor.gst_treatment
            )
            zoho_vendor.gst_no = data.get("gstn_uni", zoho_vendor.gst_no)
            zoho_vendor.pan_no = data.get("pan", zoho_vendor.pan_no)
            zoho_vendor.place_of_contact = data.get(
                "place_of_contact", zoho_vendor.place_of_contact
            )
            zoho_vendor.currency_id = data.get("currency", zoho_vendor.currency_id)
            zoho_vendor.tds_tax_id = data.get("tds", zoho_vendor.tds_tax_id)

            # Update billing address
            current_billing = zoho_vendor.billing_address or {}
            zoho_vendor.billing_address = {
                "attention": data.get(
                    "attention", current_billing.get("attention", "")
                ),
                "country": data.get("country", current_billing.get("country", "")),
                "address": data.get("address", current_billing.get("address", "")),
                "city": data.get("city", current_billing.get("city", "")),
                "state": data.get("state", current_billing.get("state", "")),
                "zip": data.get("zip_code", current_billing.get("zip", "")),
            }

            # Update shipping address
            current_shipping = zoho_vendor.shipping_address or {}
            zoho_vendor.shipping_address = {
                "attention": data.get(
                    "shipping_attention", current_shipping.get("attention", "")
                ),
                "country": data.get(
                    "shipping_country", current_shipping.get("country", "")
                ),
                "address": data.get(
                    "shipping_address", current_shipping.get("address", "")
                ),
                "city": data.get("shipping_city", current_shipping.get("city", "")),
                "state": data.get("shipping_state", current_shipping.get("state", "")),
                "zip": data.get("shipping_zip_code", current_shipping.get("zip", "")),
            }

            # Update contact persons
            contact_persons = zoho_vendor.contact_persons or []
            if contact_persons:
                # Update existing contact person
                contact_persons[0].update(
                    {
                        "first_name": data.get(
                            "first_name", contact_persons[0].get("first_name", "")
                        ),
                        "last_name": data.get(
                            "last_name", contact_persons[0].get("last_name", "")
                        ),
                        "email": data.get("email", contact_persons[0].get("email", "")),
                        "phone": data.get("phone", contact_persons[0].get("phone", "")),
                    }
                )
            else:
                # Create new contact person
                contact_persons = [
                    {
                        "first_name": data.get("first_name", ""),
                        "last_name": data.get("last_name", ""),
                        "email": data.get("email", ""),
                        "phone": data.get("phone", ""),
                    }
                ]
            zoho_vendor.contact_persons = contact_persons

            # Update bank accounts
            bank_accounts = zoho_vendor.bank_accounts or []
            if data.get("account_number"):
                bank_account = {
                    "beneficiary_name": data.get("beneficiary_name", ""),
                    "bank_name": data.get("bank_name", ""),
                    "account_number": data.get("account_number", ""),
                    "ifsc": data.get("ifsc", ""),
                }

                if bank_accounts:
                    # Update existing bank account
                    bank_accounts[0] = bank_account
                else:
                    # Add new bank account
                    bank_accounts = [bank_account]

                zoho_vendor.bank_accounts = bank_accounts

            # Update notes/remarks
            zoho_vendor.notes = data.get("remarks", zoho_vendor.notes)

            # Update last modified time
            zoho_vendor.last_modified_time = timezone.now()

            # Save ZohoVendor
            zoho_vendor.save()

            vendor.name = data.get("name", vendor.name)
            vendor.phone = data.get("phone", vendor.phone)
            vendor.hsn_or_sac = (
                int(data.get("hsn_or_sac", 0))
                if data.get("hsn_or_sac")
                else vendor.hsn_or_sac
            )

            vendor.save()

            return Response(
                {
                    "message": "Vendor updated successfully!",
                    "vendor_id": vendor.id,
                    "zoho_vendor_id": zoho_vendor.id,
                },
                status=status.HTTP_200_OK,
            )

    except Vendor.DoesNotExist:
        return Response(
            {"error": "Vendor not found"},
            status=status.HTTP_404_NOT_FOUND,
        )
    except Exception as e:
        print(f"Error updating vendor: {str(e)}")
        return Response(
            {"error": "Failed to update vendor", "details": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_client_invoices(request):
    try:
        page = request.query_params.get("page")
        salesperson_id = request.query_params.get("salesperson_id", "")
        project_id = request.query_params.get("project_id")
        project_type = request.query_params.get("project_type")
        salesorder_id = request.query_params.get("salesorder_id")
        clientinvoices = []

        if salesperson_id:
            clientinvoices = ClientInvoice.objects.filter(salesperson_id=salesperson_id)
        else:
            clientinvoices = ClientInvoice.objects.all()

        if salesorder_id:
            clientinvoices = clientinvoices.filter(
                sales_order__salesorder_id=salesorder_id
            )
        invoices = ClientInvoiceGetSerializer(
            clientinvoices,
            many=True,
        ).data
        return Response(invoices, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))

        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_all_invoices_of_sales_order(request, sales_order_id):
    try:
        sales_order_instance = SalesOrder.objects.get(salesorder_id=sales_order_id)

        client_invoices_data = ClientInvoiceSerializer(
            ClientInvoice.objects.filter(sales_order=sales_order_instance),
            many=True,
        ).data
        for invoice_data in client_invoices_data:
            # Assuming `ClientInvoiceLineItem` has a ForeignKey or reverse relationship
            line_items_data = V2ClientInvoiceLineItemSerializer(
                ClientInvoiceLineItem.objects.filter(
                    clientinvoice__invoice_id=invoice_data["invoice_id"]
                ),
                many=True,
            ).data
            invoice_data["line_items"] = line_items_data
        return Response(client_invoices_data, status=status.HTTP_200_OK)

    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_handovers_so(request, user_id, user_type):
    try:
        handovers = []
        if user_type == "pmo":
            handovers = HandoverDetails.objects.filter(
                schedular_project__isnull=True,
                caas_project__isnull=True,
                is_drafted=False,
            ).order_by("-created_at")
        else:
            handovers = HandoverDetails.objects.filter(sales__id=user_id).order_by(
                "-created_at"
            )
        all_sales_order_ids = []
        for handover in handovers:
            all_sales_order_ids.extend(handover.sales_order_ids)

        sales_orders_ids_str = ",".join(all_sales_order_ids)
        all_sales_orders = []
        if sales_orders_ids_str:
            all_sales_orders = SalesOrderGetSerializer(
                SalesOrder.objects.filter(salesorder_id__in=all_sales_order_ids),
                many=True,
            ).data
            # fetch_sales_orders(
            #     organization_id, f"&salesorder_ids={sales_orders_ids_str}"
            # )
        final_data = {}
        sales_order_dict = {
            order["salesorder_id"]: order["salesorder_number"]
            for order in all_sales_orders
        }
        if handovers:
            for handover in handovers:
                handover_id = handover.id
                res = []
                sales_order_ids = handover.sales_order_ids
                for sales_order_id in sales_order_ids:
                    sales_order_number = sales_order_dict.get(sales_order_id)
                    if sales_order_number:
                        res.append(sales_order_number)
                final_data[handover_id] = res
        return Response(final_data)

    except ObjectDoesNotExist as e:
        print(str(e))
        return Response({"error": "Project does not exist"}, status=404)

    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_total_so_created_count(request, sales_person_id):
    try:
        all_sales_orders = SalesOrderGetSerializer(
            SalesOrder.objects.filter(salesperson_id=sales_person_id), many=True
        ).data
        count = len(all_sales_orders)
        return Response(count, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_handovers_count(request, sales_person_id):
    try:
        # Count handovers where is_drafted is true
        drafted_count = HandoverDetails.objects.filter(
            sales__id=sales_person_id, is_drafted=True
        ).count()

        # Count handovers where is_accepted is true
        accepted_count = HandoverDetails.objects.filter(
            sales__id=sales_person_id, is_accepted=True
        ).count()

        # Count handovers where both is_drafted and is_accepted are false (pending)
        pending_count = HandoverDetails.objects.filter(
            sales__id=sales_person_id, is_drafted=False, is_accepted=False
        ).count()

        # Calculate total count
        total_count = drafted_count + accepted_count + pending_count

        return Response(
            {
                "drafted_count": drafted_count,
                "accepted_count": accepted_count,
                "pending_count": pending_count,
                "total_count": total_count,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(str(e))
        return Response(
            {"error": "Failed to get handover count."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def sales_orders_with_due_invoices(request, sales_person_id):
    try:
        all_sales_orders = SalesOrderGetSerializer(
            SalesOrder.objects.filter(salesperson_id=sales_person_id), many=True
        ).data
        count = len(all_sales_orders)
        return Response(count, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_line_items(request):
    current_date = datetime.now().date()
    salesperson_id = request.query_params.get("salesperson_id")
    line_items = SalesOrderLineItem.objects.filter(
        custom_field_hash__cf_due_date__isnull=False, is_invoiced=False
    )
    line_item_data = []
    for item in line_items:
        due_date = datetime.strptime(
            item.custom_field_hash["cf_due_date"], "%d/%m/%Y"
        ).date()
        if due_date <= current_date:
            sales_order = SalesOrder.objects.filter(
                so_line_items__line_item_id=item.line_item_id
            ).first()
            if sales_order and (
                not salesperson_id or sales_order.salesperson_id == salesperson_id
            ):
                line_item = {
                    "sales_order_id": sales_order.salesorder_id,
                    "sales_order_number": sales_order.salesorder_number,
                    "line_item_id": item.line_item_id,
                    "client_name": sales_order.customer_name,
                    "line_item_description": item.description,
                    "due_date": item.custom_field_hash["cf_due_date"],
                }

                if sales_order.caas_project is not None:
                    line_item["project_type"] = "Coaching"
                    line_item["project_name"] = (
                        sales_order.caas_project.name
                        if sales_order.caas_project
                        else None
                    )
                elif sales_order.schedular_project is not None:
                    line_item["project_type"] = "Skill Training"
                    line_item["project_name"] = (
                        sales_order.schedular_project.name
                        if sales_order.schedular_project
                        else None
                    )
                else:
                    line_item["project_type"] = None
                    line_item["project_name"] = None
                line_item_data.append(line_item)
    return JsonResponse(line_item_data, safe=False)


class SalesOrderLineItemListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = SalesOrderLineItemListSerializer
    pagination_class = CustomPageNumberPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = SalesOrderLineItemFilter

    def get_queryset(self):
        salesperson_id = self.request.query_params.get("salesperson_id")

        # Filter and annotate
        line_items = SalesOrderLineItem.objects.annotate(
            due_date=salesorder_line_item_due_date_case
        ).filter(due_date__isnull=False, is_invoiced=False, salesorder__isnull=False)

        if salesperson_id:
            # Filter line items by salesperson via the related sales order
            line_items = line_items.filter(
                salesorder__in=SalesOrder.objects.filter(salesperson_id=salesperson_id)
            )

        return line_items.order_by("-due_date").prefetch_related("salesorder_set")

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Calculate the sum of the `item_total` for all filtered entries
        total_sum = (
            queryset.annotate(
                calculated_total=F("item_total") * F("salesorder__exchange_rate")
            ).aggregate(total_sum=Sum("calculated_total"))["total_sum"]
            or 0.0
        )

        # Paginate the queryset
        page = self.paginate_queryset(queryset)

        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(
                {"line_items": serializer.data, "total_sum": total_sum}
            )

        serializer = self.get_serializer(queryset, many=True)
        return Response({"line_items": serializer.data, "total_sum": total_sum})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_ctt_revenue_data(request):
    start_date = request.query_params.get("start_date", "")
    end_date = request.query_params.get("end_date", "")
    revenue_data = get_revenue_data("CTT", start_date, end_date)
    status_code = (
        status.HTTP_200_OK
        if "result" in revenue_data
        else status.HTTP_500_INTERNAL_SERVER_ERROR
    )
    return Response(revenue_data, status=status_code)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_meeraq_revenue_data(request):
    start_date = request.query_params.get("start_date", "")
    end_date = request.query_params.get("end_date", "")
    revenue_data = get_revenue_data("Meeraq", start_date, end_date)
    status_code = (
        status.HTTP_200_OK
        if "result" in revenue_data
        else status.HTTP_500_INTERNAL_SERVER_ERROR
    )
    return Response(revenue_data, status=status_code)


@api_view(["GET"])
@permission_classes([AllowAny])
def get_line_items_detail_in_excel(request):
    line_items = SalesOrderLineItem.objects.filter(salesorder__isnull=False)
    data = line_items.values(
        "salesorder__salesorder_number",
        "description",
        "custom_field_hash__cf_due_date",
        "rate",
        "quantity",
        "quantity_invoiced",
        "tax_percentage",
        "item_total",
    )
    df = pd.DataFrame(data)
    df = df.rename(
        columns={
            "salesorder__salesorder_number": "Sales Order Number",
            "description": "Items and description",
            "custom_field_hash__cf_due_date": "Due Date",
            "rate": "Rate",
            "quantity": "Quantity",
            "quantity_invoiced": "Quantity Invoiced",
            "tax_percentage": "Tax Percentage",
            "item_total": "Item Total",
        }
    )
    # Fill empty due dates with hyphen
    df["Due Date"] = df["Due Date"].fillna("-")

    # Convert Quantity and Quantity Invoiced columns to floating-point numbers
    df["Quantity"] = df["Quantity"].astype(float)
    df["Quantity Invoiced"] = df["Quantity Invoiced"].astype(float)

    # Round the floating-point numbers to integers
    df["Quantity"] = df["Quantity"].round(0).astype(int)
    df["Quantity Invoiced"] = df["Quantity Invoiced"].round(0).astype(int)

    # Save the DataFrame to an Excel file in-memory
    excel_data = BytesIO()
    df.to_excel(excel_data, index=False)
    excel_data.seek(0)

    # Create the response with the Excel file
    response = HttpResponse(
        excel_data.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename=line_items_details.xlsx"

    return response


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def update_vendor_msme(request, vendor_id):
    vendor = Vendor.objects.get(id=vendor_id)
    vendor.is_msme = request.data.get("is_msme", None)
    vendor.save()
    return Response({"message": "MSME status updated successfully!"})


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsInRoles("pmo", "finance")])
def edit_purchase_order(request, po_id):
    try:
        purchase_order_data = PurchaseOrder.objects.filter(
            purchaseorder_id=po_id
        ).first()
        if not purchase_order_data:
            return None, status.HTTP_404_NOT_FOUND, "Purchase order not found"
        data = json.loads(request.data.get("JSONString"))

        line_items_data = process_line_item_custom_fields(data["line_items"])
        line_items_data = process_po_line_item_data(line_items_data, True)
        vendor = ZohoVendor.objects.get(contact_id=purchase_order_data.vendor_id)
        purchase_order_data.zoho_vendor = vendor
        purchase_order_data.save()
        # Check PMO brand for guest CTT flag
        rate = get_exchange_rate(vendor.currency_code, "INR")
        print("Step 1")
        line_item_totals = get_purchase_order_totals(line_items_data)
        print("Step 2")
        # Associate line items with the sales order
        purchase_order_data.total = line_item_totals["total"]
        purchase_order_data.sub_total = line_item_totals["sub_total"]
        purchase_order_data.sub_total_inclusive_of_tax = line_item_totals[
            "sub_total_inclusive_of_tax"
        ]
        purchase_order_data.tax_total = line_item_totals["tax_total"]
        purchase_order_data.total_quantity = line_item_totals["total_quantity"]

        purchase_order_data.last_modified_time = datetime.now()
        purchase_order_data.exchange_rate = rate

        purchase_order_data.save()
        # Process line items

        for index, line_item in enumerate(line_items_data):
            print(f"DEBUG: Processing line item {index + 1}")
            line_item_instance = PurchaseOrderLineItem.objects.filter(
                line_item_id=line_item["line_item_id"]
            ).first()
            line_item_serializer = None
            if line_item_instance:
                line_item_serializer = PurchaseOrderLineItemSerializer(
                    line_item_instance, data=line_item, partial=True
                )
            else:
                line_item_serializer = PurchaseOrderLineItemSerializer(data=line_item)

            if line_item_serializer.is_valid():
                print(f"DEBUG: Line item {index + 1} valid, saving")
                instance = line_item_serializer.save()
                purchase_order_data.po_line_items.add(instance)
            else:
                print(
                    f"DEBUG: Line item {index + 1} invalid for PO {data['purchaseorder_number']}"
                    f"Errors: {line_item_serializer.errors}"
                )
        return Response({"message": "Purchase Order created successfully."})

    except Exception as e:
        print(str(e))
        return Response({"error": "Failed to update purchase order."}, status=500)


class EntityListCreateView(generics.ListCreateAPIView):
    queryset = Entity.objects.all()
    serializer_class = EntitySerializer


# Retrieve, Update, and Delete View
class EntityRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Entity.objects.all()
    serializer_class = EntitySerializer


class ZohoCustomerListCreateView(generics.ListCreateAPIView):
    queryset = ZohoCustomer.objects.all().order_by("-created_time")
    serializer_class = ZohoCustomerSerializer
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ["contact_name", "company_name", "email"]

    def perform_create(self, serializer):
        instance = serializer.save()
        existing_custom_field_hash = instance.custom_field_hash or {}
        instance.custom_field_hash = existing_custom_field_hash
        instance.save()
        return instance

    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)


class ZohoCustomerUpdateView(generics.UpdateAPIView):
    queryset = ZohoCustomer.objects.all()
    serializer_class = ZohoCustomerSerializer

    def perform_update(self, serializer):
        instance = self.get_object()
        instance = serializer.save()
        existing_custom_field_hash = instance.custom_field_hash or {}
        instance.custom_field_hash = existing_custom_field_hash
        instance.save()
        return instance

    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)


class CompanyListCreateView(generics.ListCreateAPIView):
    queryset = Company.objects.filter(is_deleted=False).order_by("-created_time")
    serializer_class = CompanySerializer
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    filterset_class = CompanyFilter
    search_fields = [
        "account_name",
        "phone",
        "billing_street",
        "billing_country",
        "owner__email",
        "billing_state",
        "website",
        "industry",
        "owner",
    ]

    def get_queryset(self):
        user_id = self.request.query_params.get("user_id")
        user_type = self.request.query_params.get("usertype")
        comapny_filter = Q(is_deleted=False)
        if user_id == "all":
            comapny_filter = Q(is_deleted=False)
        else:
            comapny_filter = (
                Q(added_by__id=user_id, is_deleted=False)
                if user_id
                else Q(added_by=self.request.user, is_deleted=False)
            )
            role = Role.objects.get(name=user_type)
            filter_subordinates = get_subordinates_of_a_user_in_role(
                "Company", self.request.user, role
            )
            if filter_subordinates:
                comapny_filter |= filter_subordinates
        queryset = Company.objects.filter(comapny_filter).order_by("-created_time")
        return queryset

    def post(self, request, *args, **kwargs):
        # Validate the incoming data using the serializer
        data = {
            "data": [
                {
                    **request.data,
                    "Owner": get_owner_details(),
                    "Tag": [],
                }
            ]
        }
        try:
            response = zoho_api_request("POST", "Accounts", data)

            create_or_update_company(response["data"][0]["details"]["id"], request)
        except requests.HTTPError as e:
            return Response(
                {"error": str(e), "details": e.response.json()},
                status=e.response.status_code,
            )

        # Save to local database only if Zoho creation was successful

        return Response(
            {"message": "Company created successfully."}, status=status.HTTP_201_CREATED
        )


class CompanyDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Company.objects.filter(is_deleted=False).order_by("-created_time")
    serializer_class = CompanySerializer
    lookup_field = "account_id"

    def put(self, request, *args, **kwargs):
        # Validate the incoming data using the serializer
        data = {
            "data": [
                {
                    **request.data,
                    "Owner": get_owner_details(),
                    "Tag": [],
                }
            ]
        }
        try:
            response = zoho_api_request(
                "PUT", f"Accounts/{self.get_object().account_id}", data
            )

            create_or_update_company(response["data"][0]["details"]["id"], request)
        except requests.HTTPError as e:
            return Response(
                {"error": str(e), "details": e.response.json()},
                status=e.response.status_code,
            )

        # Save to local database only if Zoho creation was successful

        return Response(
            {"message": "Company updated successfully."}, status=status.HTTP_200_OK
        )

    def perform_destroy(self, instance):
        # Mark the local record as deleted in any scenario
        instance.is_deleted = True

        try:
            # Attempt to delete from Zoho
            zoho_api_request("DELETE", f"Accounts/{instance.account_id}")
        except requests.HTTPError as e:
            # Log or print the error details
            # Be aware that e.response.json() may fail if the response isn't valid JSON
            print({"error": str(e), "details": e.response.json()})

            # Raise a DRF-friendly exception
            raise Exception(
                {
                    "message": "Company deleted locally, but could not be deleted on Zoho.",
                    "error": str(e),
                }
            )
        finally:
            # Always save the updated 'is_deleted' state
            instance.save()


# Contact Views
class ContactListCreateView(generics.ListCreateAPIView):
    queryset = Contact.objects.filter(is_deleted=False).order_by("-created_time")
    serializer_class = ContactSerializer
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    filterset_class = ContactFilter
    search_fields = [
        "first_name",
        "last_name",
        "full_name",
        "title",
        "email",
        "phone",
        "mailing_city",
        "mailing_state",
        "mailing_country",
        "account_name__name",
    ]

    def get_queryset(self):
        user_id = self.request.query_params.get("user_id")
        user_type = self.request.query_params.get("usertype")
        contact_filter = Q(is_deleted=False)
        if user_id == "all":
            contact_filter = Q(is_deleted=False)
        else:
            contact_filter = (
                Q(added_by__id=user_id, is_deleted=False)
                if user_id
                else Q(added_by=self.request.user, is_deleted=False)
            )
            role = Role.objects.get(name=user_type)
            filter_subordinates = get_subordinates_of_a_user_in_role(
                "Contact", self.request.user, role
            )
            if filter_subordinates:
                contact_filter |= filter_subordinates

        queryset = Contact.objects.filter(contact_filter).order_by("-created_time")
        return queryset

    def post(self, request, *args, **kwargs):
        # Validate the incoming data using the serializer
        company = Company.objects.filter(id=request.data["Account_Name"]).first()

        data = {
            "data": [
                {
                    **request.data,
                    "Owner": get_owner_details(),
                    "Account_Name": {"id": company.account_id},
                }
            ]
        }

        # Attempt to create the company in Zoho

        try:
            response = zoho_api_request("POST", "Contacts", data)

            create_or_update_contacts(response["data"][0]["details"]["id"], request)
        except requests.HTTPError as e:
            return Response(
                {"error": str(e), "details": e.response.json()},
                status=e.response.status_code,
            )

        # Save to local database only if Zoho creation was successful

        return Response(
            {"message": "Contact created successfully."}, status=status.HTTP_201_CREATED
        )


class ContactDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Contact.objects.filter(is_deleted=False).order_by("-created_time")
    serializer_class = ContactSerializer
    lookup_field = "contact_id"

    def put(self, request, *args, **kwargs):
        # Validate the incoming data using the serializer
        company = Company.objects.filter(
            account_id=request.data["Account_Name"]
        ).first()

        data = {
            "data": [
                {
                    **request.data,
                    "Owner": get_owner_details(),
                    "Account_Name": {"id": company.account_id},
                }
            ]
        }

        # Attempt to create the company in Zoho

        try:
            response = zoho_api_request(
                "PUT", f"Contacts/{self.get_object().contact_id}", data
            )

            create_or_update_contacts(response["data"][0]["details"]["id"], request)
        except requests.HTTPError as e:
            return Response(
                {"error": str(e), "details": e.response.json()},
                status=e.response.status_code,
            )

        # Save to local database only if Zoho creation was successful

        return Response(
            {"message": "Contact updated successfully."}, status=status.HTTP_200_OK
        )

    def perform_destroy(self, instance):
        # Mark the local record as deleted in any scenario
        instance.is_deleted = True
        try:
            # Attempt to delete from Zoho
            zoho_api_request("DELETE", f"Contacts/{instance.contact_id}")
        except requests.HTTPError as e:
            # Log or print the error details
            # Be aware that e.response.json() may fail if the response isn't valid JSON
            print({"error": str(e), "details": e.response.json()})

            # Raise a DRF-friendly exception
            raise Exception(
                {
                    "message": "Contact deleted locally, but could not be deleted on Zoho.",
                    "error": str(e),
                }
            )
        finally:
            # Always save the updated 'is_deleted' state
            instance.save()


# Deal Views
class DealListCreateView(generics.ListCreateAPIView):
    queryset = Deal.objects.filter(is_deleted=False).order_by("-created_time")
    serializer_class = DealSerializerDepthOne
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = [
        "deal_name",
        "amount",
        "stage",
        "contact_name__name",
        "account_name__name",
    ]

    def get_queryset(self):
        user_id = self.request.query_params.get("user_id")
        user_type = self.request.query_params.get("usertype")
        deals_filter = Q(is_deleted=False)
        if user_id == "all":
            deals_filter = Q(is_deleted=False)
        else:
            deals_filter = (
                Q(added_by__id=user_id, is_deleted=False)
                if user_id
                else Q(added_by=self.request.user, is_deleted=False)
            )
            role = Role.objects.get(name=user_type)
            filter_subordinates = get_subordinates_of_a_user_in_role(
                "Deal", self.request.user, role
            )
            if filter_subordinates:
                deals_filter |= filter_subordinates
        queryset = Deal.objects.filter(deals_filter).order_by("-created_time")
        return queryset

    def post(self, request, *args, **kwargs):
        # Validate the incoming data using the serializer
        company = Company.objects.filter(id=request.data["Account_Name"]).first()
        contact = Contact.objects.filter(id=request.data["Contact_Name"]).first()

        data = {
            "data": [
                {
                    **request.data,
                    "Owner": get_owner_details(),
                    "Account_Name": {"id": company.account_id},
                    "Contact_Name": {"id": contact.contact_id},
                }
            ]
        }

        try:
            response = zoho_api_request("POST", "Deals", data, True)

            deal = create_or_update_deals(response["data"][0]["details"]["id"])

            if request.data["Stage"] == "Proposal/Price Quote":
                gm_sheet = GmSheet.objects.filter(id=request.data["gm_sheet"]).first()
                deal.gm_sheet = gm_sheet
                deal.sales_order = None
                deal.save()
            elif request.data["Stage"] == "Closed Won":
                sales_order = SalesOrder.objects.filter(
                    id=request.data["sales_order"]
                ).first()
                deal.sales_order = sales_order
                deal.save()
        except requests.HTTPError as e:
            return Response(
                {"error": str(e), "details": e.response.json()},
                status=e.response.status_code,
            )
        return Response(
            {"message": "Deal created successfully."}, status=status.HTTP_201_CREATED
        )


class DealDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Deal.objects.filter(is_deleted=False).order_by("-created_time")
    serializer_class = DealSerializer
    lookup_field = "deal_id"

    def put(self, request, *args, **kwargs):
        # Validate the incoming data using the serializer
        company = Company.objects.filter(id=request.data["Account_Name"]).first()
        contact = Contact.objects.filter(id=request.data["Contact_Name"]).first()

        data = {
            "data": [
                {
                    **request.data,
                    "Owner": get_owner_details(),
                    "Account_Name": {"id": company.account_id},
                    "Contact_Name": {"id": contact.contact_id},
                }
            ]
        }

        try:

            response = zoho_api_request(
                "PUT", f"Deals/{self.get_object().deal_id}", data, True
            )

            deal = create_or_update_deals(response["data"][0]["details"]["id"])
            if request.data["Stage"] == "Proposal/Price Quote":
                gm_sheet = GmSheet.objects.filter(id=request.data["gm_sheet"]).first()
                deal.gm_sheet = gm_sheet

                deal.save()
            elif request.data["Stage"] == "Closed Won":
                sales_order = SalesOrder.objects.filter(
                    id=request.data["sales_order"]
                ).first()
                deal.sales_order = sales_order

                deal.save()
        except requests.HTTPError as e:
            return Response(
                {"error": str(e), "details": e.response.json()},
                status=e.response.status_code,
            )
        return Response(
            {"message": "Deal updated successfully."}, status=status.HTTP_200_OK
        )

    def perform_destroy(self, instance):
        # Mark the local record as deleted in any scenario
        instance.is_deleted = True

        try:
            # Attempt to delete from Zoho
            zoho_api_request("DELETE", f"Deals/{instance.deal_id}")
        except requests.HTTPError as e:
            # Log or print the error details
            # Be aware that e.response.json() may fail if the response isn't valid JSON
            print({"error": str(e), "details": e.response.json()})

            # Raise a DRF-friendly exception
            raise Exception(
                {
                    "message": "Deal deleted locally, but could not be deleted on Zoho.",
                    "error": str(e),
                }
            )
        finally:
            # Always save the updated 'is_deleted' state
            instance.save()


class DealFileUploadView(generics.GenericAPIView):
    queryset = Deal.objects.filter(is_deleted=False)
    serializer_class = (
        DealSerializer  # or create a dedicated serializer for the file if needed
    )
    lookup_field = "deal_id"
    # parser_classes = [MultiPartParser, FormParser]  # enable file upload parsing

    def put(self, request, *args, **kwargs):
        print("hello")
        deal = get_object_or_404(Deal, deal_id=kwargs.get("deal_id"), is_deleted=False)

        # Expecting the file field name in the request to be 'file'
        uploaded_file = request.data.get("deal_file")
        if not uploaded_file:
            return Response(
                {"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST
            )

        # Assuming your Deal model has a FileField or similar attribute called 'attachment' or 'file'
        # Replace 'attachment' below with the actual field name in your model
        deal.deal_file = uploaded_file
        deal.save()

        return Response(
            {"message": "File uploaded successfully."}, status=status.HTTP_200_OK
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_pipelines_data(request):
    try:
        response = zoho_api_request("GET", "settings/pipelines")
        return Response(response, status=status.HTTP_200_OK)
    except Exception as e:
        print(str(e))
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class BankDetailsListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    queryset = BankDetails.objects.all()
    serializer_class = BankDetailsSerializer


class PaymentListCreateView(SerializerByMethodMixin, generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Payment.objects.all()
    serializer_map = {
        "GET": PaymentSerializer,
        "POST": PaymentCreateSerializer,
    }
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = PaymentFilters


class PaymentsExportView(PaymentListCreateView):
    pagination_class = None

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Initialize Excel workbook and worksheet
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Payments")

        # Define styles
        styles = get_styles(workbook)

        # Headers and column widths
        headers, column_widths = self.get_headers_and_widths()
        self.write_headers(worksheet, headers, styles["header_style"], column_widths)

        # Write data rows
        self.write_data_rows(worksheet, queryset, styles)

        # Finalize workbook
        worksheet.autofilter(0, 0, len(queryset), len(headers) - 1)
        worksheet.freeze_panes(1, 0)
        workbook.close()

        # Prepare the response
        output.seek(0)
        filename = f'payments_{datetime.now().strftime("%d%m%Y")}.xlsx'
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @staticmethod
    def get_headers_and_widths():
        headers = [
            "Payment Number",
            "Customer Name",
            "Invoice Number",
            "Deposit To",
            "Amount",
            "Bank Charges",
            "Currency Code",
            "Currency Symbol",
            "Exchange Rate",
            "Reference Number",
            "Date",
            "Payment Mode",
            "Created Time",
        ]
        column_widths = {
            0: 20,  # Payment Number
            1: 30,  # Customer Name
            2: 20,  # Invoice Number
            3: 20,  # Deposit To
            4: 15,  # Amount
            5: 15,  # Bank Charges
            6: 15,  # Currency Code
            7: 10,  # Currency Symbol
            8: 15,  # Exchange Rate
            9: 20,  # Reference Number
            10: 15,  # Date
            11: 15,  # Payment Mode
            12: 20,  # Created Time
        }
        return headers, column_widths

    @staticmethod
    def write_headers(worksheet, headers, header_style, column_widths):
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_style)
            worksheet.set_column(col, col, column_widths.get(col, 15))

    def write_data_rows(self, worksheet, queryset, styles):
        for row, payment in enumerate(queryset, start=1):
            worksheet.write(row, 0, payment.payment_number, styles["cell_style"])
            worksheet.write(
                row,
                1,
                payment.customer.contact_name if payment.customer else "",
                styles["cell_style"],
            )
            worksheet.write(
                row,
                2,
                payment.client_invoice.invoice_number if payment.client_invoice else "",
                styles["cell_style"],
            )
            worksheet.write(
                row,
                3,
                payment.deposit_to.display_name if payment.deposit_to else "",
                styles["cell_style"],
            )
            worksheet.write(row, 4, float(payment.amount), styles["number_style"])
            worksheet.write(row, 5, float(payment.bank_charges), styles["number_style"])
            worksheet.write(row, 6, payment.currency_code, styles["cell_style"])
            worksheet.write(row, 7, payment.currency_symbol, styles["cell_style"])
            worksheet.write(row, 8, payment.exchange_rate or "", styles["cell_style"])
            worksheet.write(row, 9, payment.reference_number, styles["cell_style"])
            worksheet.write(row, 10, payment.date, styles["date_style"])
            worksheet.write(row, 11, payment.payment_mode, styles["cell_style"])
            worksheet.write(
                row,
                12,
                payment.created_time.strftime("%Y-%m-%d %H:%M:%S"),
                styles["cell_style"],
            )


class PaymentUpdateView(generics.UpdateAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Payment.objects.all()
    serializer_class = PaymentCreateSerializer


class DownloadPaymentReceipt(APIView):
    permission_classes = [IsAuthenticated, IsInRoles("pmo", "vendor", "finance")]

    def get(self, request, id):
        try:
            payment = get_object_or_404(Payment, id=id)
            amount_to_words = amount_convert_to_words(
                Decimal(payment.amount), payment.currency_code
            )
            logo = "templates/logos/invoice_logo.png"
            image_base64 = None
            if logo:
                with open(logo, "rb") as image_file:
                    image_base64 = base64.b64encode(image_file.read()).decode("utf-8")
            signature = "templates/logos/rajat_sign.png"
            if signature:
                with open(signature, "rb") as image_file:
                    signature_image_base64 = base64.b64encode(image_file.read()).decode(
                        "utf-8"
                    )
            data = {
                "image_base64": image_base64,
                "organization_name": (
                    payment.client_invoice.sales_order.entity.name
                    if payment.client_invoice
                    else payment.bill.entity.name
                ),
                "organization_billing_address": (
                    payment.client_invoice.sales_order.entity.billing_address
                    if payment.client_invoice
                    else payment.bill.entity.billing_address
                ),
                "payment_number": payment.payment_number,
                "payment_date": payment.date.strftime("%d/%m/%Y"),
                "customer_name": (
                    payment.customer.contact_name
                    if payment.customer
                    else payment.bill.zoho_vendor.contact_name
                ),
                "customer_address": (
                    format_address(payment.customer.billing_address)
                    if payment.customer
                    else format_address(payment.bill.zoho_vendor.billing_address)
                ),
                "currency_symbol": payment.currency_symbol or "N/A",
                "amount": payment.amount,
                "payment_mode": dict(payment.PAYMENT_MODES).get(
                    payment.payment_mode, "N/A"
                ),
                "reference_number": payment.reference_number or "N/A",
                "amount_in_words": amount_to_words.title(),  # Helper method to convert to words
                "signature_image_base64": signature_image_base64,
                "invoice_number": (
                    payment.client_invoice.invoice_number
                    if payment.client_invoice
                    else payment.bill.bill_number
                ),
                "invoice_amount": f"{payment.client_invoice.currency_symbol if payment.client_invoice else payment.bill.currency_symbol} {payment.client_invoice.total if payment.client_invoice else payment.bill.total}",
                "invoice_date": (
                    payment.client_invoice.date.strftime("%d/%m/%Y")
                    if payment.client_invoice
                    else payment.bill.date
                ),
                "payment_amount": f"{payment.currency_symbol or ''} {payment.amount}",
                "type": payment.type,
                "bank_name": (
                    payment.deposit_to.bank_name
                    if payment.client_invoice
                    else payment.debited_from.bank_name
                ),
            }
            email_message = render_to_string(
                "finance/payment_receipt.html",
                data,
            )
            pdf = pdfkit.from_string(email_message, False, configuration=pdfkit_config)
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = (
                f'attachment; filename={f"{payment.payment_number}_payment.pdf"}'
            )
            return response

        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to download invoice."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class FilteredPurchaseOrderView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = OPEPurchaseOrderSerializer
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    filterset_class = OpePurchaseOrderFilter
    search_fields = [
        "purchaseorder_number",
        "vendor_name",
        "status",
        "clientinvoice__status",
        "clientinvoice__invoice_number",
        "caas_project__name",
        "schedular_project__name",
    ]

    def get_queryset(
        self,
    ):
        invoice_pending = self.request.query_params.get("invoice_pending")
        status = self.request.query_params.get("status_client_invoice")

        queryset = PurchaseOrder.objects.filter(
            is_billed_to_client=True,
        )

        if invoice_pending:
            queryset = queryset.filter(clientinvoice__isnull=True)
        if status:
            if status == "paid":
                queryset = queryset.filter(clientinvoice__status="paid")
            else:
                queryset = queryset.filter(
                    Q(clientinvoice__isnull=True) | ~Q(clientinvoice__status="paid")
                )

        return queryset


class FilteredPurchaseOrderViewExcel(generics.ListAPIView):
    permission_classes = [AllowAny]
    serializer_class = PurchaseOrderSerializer

    def get_queryset(self):
        return PurchaseOrder.objects.filter(is_billed_to_client=True)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        data = []

        for po in queryset:

            client_invoice = ClientInvoice.objects.filter(purchase_orders=po).first()
            client_invoice_data = (
                {
                    "invoice_number": (
                        client_invoice.invoice_number if client_invoice else None
                    ),
                    "total": client_invoice.total if client_invoice else None,
                    "status": client_invoice.status if client_invoice else None,
                    "currency_symbol": (
                        client_invoice.currency_symbol if client_invoice else None
                    ),
                    "invoiced_date": client_invoice.date if client_invoice else None,
                    "payment_date": (
                        client_invoice.last_payment_date if client_invoice else None
                    ),
                }
                if client_invoice
                else None
            )

            entry = {
                "purchaseorder_id": po.purchaseorder_id,
                "purchase_order_no": po.purchaseorder_number,
                "facilitator_name": po.vendor_name,
                "raised_amount": {
                    "currency_symbol": po.currency_symbol,
                    "total": po.total,
                },
                "po_status": po.status,
                "client_invoice": client_invoice_data,
            }
            data.append(entry)

        return Response(data)

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Purchase Orders"

        # Define headers
        headers = [
            "Purchase Order No.",
            "Purchase Order Id",
            "Facilitator Name",
            "Raised Amount",
            "Client Invoice",
            "PO Status",
        ]
        sheet.append(headers)

        # Add data rows
        for po in queryset:

            client_invoice = ClientInvoice.objects.filter(purchase_order=po).first()
            client_invoice_number = (
                client_invoice.invoice_number if client_invoice else "N/A"
            )

            row = [
                po.purchaseorder_number,
                po.purchaseorder_id,
                po.vendor_name or "N/A",
                f"{po.currency_symbol} {po.total}",
                client_invoice_number,
                po.status,
            ]
            sheet.append(row)

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=PurchaseOrders.xlsx"
        workbook.save(response)
        return response


class CreditNoteListCreateView(generics.ListCreateAPIView):
    queryset = CreditNote.objects.all()
    serializer_class = CreditNoteSerializer
    permission_classes = [IsAuthenticated, IsInRoles("pmo", "finance", "sales")]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = [
        "brand",
        "customer",
        "credit_note_number",
        "client_invoice__invoice_number",
        "customer__contact_name",
    ]
    search_fields = [
        "credit_note_number",
        "customer_notes",
        "client_invoice__invoice_number",
        "customer__contact_name",
    ]
    # ordering_fields = ['created_time', 'updated_time', 'amount']

    def perform_create(self, serializer):
        # Additional creation logic can be added here
        credit_note = serializer.save()
        if credit_note.client_invoice:
            update_invoice_status_and_balance(credit_note.client_invoice)


class CreditNoteDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CreditNote.objects.all()
    serializer_class = CreditNoteSerializer
    permission_classes = [IsAuthenticated, IsInRoles("pmo", "finance", "sales")]

    def perform_update(self, serializer):
        # Additional update logic can be added here
        credit_note = serializer.save()
        if credit_note.client_invoice:
            update_invoice_status_and_balance(credit_note.client_invoice)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        client_invoice = instance.client_invoice
        self.perform_destroy(instance)
        if client_invoice:
            update_invoice_status_and_balance(client_invoice)
        return Response(
            {"message": "Credit note deleted successfully"},
            status=status.HTTP_204_NO_CONTENT,
        )


class CreditNoteView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, credit_note_id):
        try:
            credit_note = CreditNote.objects.get(id=credit_note_id)
            invoice = credit_note.client_invoice
            bank_detail = (
                BankDetailsSerializer(invoice.bank_detail).data
                if invoice.bank_detail
                else {}
            )

            amount_to_words = amount_convert_to_words(
                Decimal(invoice.total), invoice.currency_code
            )
            entity = Entity.objects.exclude(id=int(env("INDIA_ENTITY_ID"))).first()

            data = {
                "organization_name": entity.name,
                "organization_billing_address": entity.billing_address,
                "customer_name": credit_note.customer.contact_name,
                "customer_address": format_address(
                    credit_note.customer.billing_address
                ),
                "customer_email": credit_note.customer.email,
                "customer_phone": credit_note.customer.phone,
                "invoice_number": credit_note.credit_note_number,
                "customer_notes": credit_note.customer_notes,
                "invoice_date": (
                    credit_note.created_at.strftime("%d/%m/%Y")
                    if credit_note.created_at
                    else ""
                ),
                "sales_order_id": invoice.sales_order.salesorder_id,
                "sales_order_no": invoice.sales_order.salesorder_number,
                "currency_code": invoice.currency_code,
                "currency_symbol": invoice.currency_symbol,
                "line_items": get_Client_invoice_line_items(
                    invoice.client_invoice_line_items.all()
                ),
                "total": invoice.total,
                "amount_to_words": amount_to_words,
            }

            # Render HTML template
            html_content = render_to_string("creditnote.html", data)

            # Generate PDF
            pdf = pdfkit.from_string(html_content, False, configuration=pdfkit_config)

            # Create response
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = (
                f'attachment; filename="Credit_Note_{credit_note.credit_note_number}.pdf"'
            )

            return response
        except Exception as e:
            print(str(e))
            return Response(
                {"error": "Failed to download credit note."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class DeleteCustomerAPIView(generics.DestroyAPIView):
    queryset = ZohoCustomer.objects.all()
    lookup_url_kwarg = "customer_id"


class DownloadCreditNotesAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        # Fetch all CreditNote instances
        credit_notes = CreditNote.objects.select_related("customer", "client_invoice")

        # Create a new workbook and set up the sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Credit Notes"

        # Define headers
        headers = [
            "Brand",
            "Customer Name",
            "Credit Note Number",
            "Client Invoice Number",
            "Amount",
            "Terms and Conditions",
            "Customer Notes",
        ]
        sheet.append(headers)

        # Populate data rows
        for credit_note in credit_notes:
            print("credit")
            sheet.append(
                [
                    credit_note.brand or "N/A",
                    (
                        credit_note.customer.first_name + credit_note.customer.last_name
                        if credit_note.customer
                        else "N/A"
                    ),
                    credit_note.credit_note_number or "N/A",
                    (
                        credit_note.client_invoice.invoice_number
                        if credit_note.client_invoice
                        else "N/A"
                    ),
                    credit_note.client_invoice.total or "N/A",
                    credit_note.terms_and_conditions or "N/A",
                    credit_note.customer_notes or "N/A",
                ]
            )

        # Create the HTTP response with Excel content
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Credit_Notes.xlsx"'
        workbook.save(response)

        return response


class BillListCreateAPIView(generics.ListCreateAPIView):
    serializer_class = V2BillSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ["bill_number", "zoho_vendor__contact_name"]

    def get_queryset(self):
        """
        Optionally restricts the returned bills by filtering against
        query parameters in the URL.
        """
        entity = Entity.objects.last()

        queryset = Bill.objects.filter(entity=entity)

        # Filter by vendor
        vendor_id = self.request.query_params.get("vendor_id", None)
        if vendor_id:
            queryset = queryset.filter(zoho_vendor_id=vendor_id)

        # Filter by date range
        start_date = self.request.query_params.get("start_date", None)
        end_date = self.request.query_params.get("end_date", None)
        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])

        # Filter by status
        status = self.request.query_params.get("status", None)
        if status:
            queryset = queryset.filter(status=status)

        return queryset

    def create(self, request, *args, **kwargs):
        data = request.data
        purchase_orders = None
        invoice_number = request.data["invoice_number"]
        vendor = ZohoVendor.objects.get(contact_id=request.data["vendor_id"])
        data["zoho_vendor"] = vendor.id
        purchase_order_ids = request.data["purchaseorder_ids"]
        entity = Entity.objects.last()
        data["entity"] = entity.id
        line_items = data.pop("bill_line_items", [])

        # Remove purchase_orders from data since we'll set it after creation
        if "purchase_orders" in data:
            data.pop("purchase_orders")

        serializer = self.get_serializer(data=data, context={"line_items": line_items})
        serializer.is_valid(raise_exception=True)

        # Save the instance
        instance = serializer.save()

        # Set the many-to-many relationship after instance is created
        if purchase_order_ids:
            purchase_orders = PurchaseOrder.objects.filter(
                purchaseorder_id__in=purchase_order_ids
            )
            instance.purchase_orders.set(purchase_orders)
        for purchase_order in purchase_orders:
            quantiy_invocied_in_items = get_po_quantity_invoices(
                instance.bill_line_items.all()
            )
            # updating so quantity input and if so line item is invoiced
            for line_item in purchase_order.po_line_items.all():
                if line_item.line_item_id in quantiy_invocied_in_items:

                    line_item.quantity_billed += quantiy_invocied_in_items[
                        line_item.line_item_id
                    ]
                    line_item.save()

            po_line_item_totals = get_purchase_order_instance_totals(
                purchase_order.po_line_items.all()
            )
            if (
                po_line_item_totals["total_quantity"]
                == po_line_item_totals["total_invoiced_quantity"]
            ):
                purchase_order.status = "billed"
                purchase_order.billed_status = "billed"
            else:
                purchase_order.status = "partially_billed"
                purchase_order.billed_status = "partially_billed"

            purchase_order.save()
        # adding bill to invoice data
        map_bill_to_invoice(instance, invoice_number)
        headers = self.get_success_headers(serializer.data)
        return Response(
            serializer.data, status=status.HTTP_201_CREATED, headers=headers
        )


class BillRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Bill.objects.all()
    serializer_class = V2BillSerializer
    permission_classes = [IsAuthenticated]

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        line_items = request.data.pop("bill_line_items", [])
        serializer = self.get_serializer(
            instance,
            data=request.data,
            context={"line_items": line_items},
            partial=partial,
        )
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        # Adding bill to invoice data
        map_bill_to_invoice(instance)
        return Response(serializer.data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def generate_purchase_order_pdf(request, purchase_order_id):
    """
    Generate PDF for a specific purchase order
    """
    # Get the purchase order instance
    po = get_object_or_404(PurchaseOrder, purchaseorder_id=purchase_order_id)

    # Get line items
    line_items = po.po_line_items.all().order_by("item_order")

    # Get vendor details
    vendor = po.zoho_vendor
    entity = Entity.objects.last()
    context = {
        "logo_path": "https://blueskycsr.com/wp-content/uploads/2023/10/Bluesky-Logo.png",
        # Company Information
        "company_info": {
            "name": entity.name,  # You might want to store this in settings
            "address_line": entity.billing_address,
        },
        # Vendor Information
        "vendor_info": {
            "name": vendor.contact_name if vendor else po.vendor_name,
            "address": (
                po.billing_address.get("address", "") if po.billing_address else ""
            ),
            "city": po.billing_address.get("city", "") if po.billing_address else "",
            "state": po.billing_address.get("state", "") if po.billing_address else "",
            "country": (
                po.billing_address.get("country", "") if po.billing_address else ""
            ),
            "zip": po.billing_address.get("zip", "") if po.billing_address else "",
        },
        # PO Details
        "po_details": {
            "number": po.purchaseorder_number,
            "date": po.date,
            "reference": po.reference_number,
            "currency_symbol": po.currency_symbol or "₹",
            "payment_terms": po.payment_terms_label,
        },
        # Line Items
        "items": [
            {
                "id": index + 1,
                "description": item.description,
                "quantity": item.quantity,
                "rate": item.rate,
                "amount": item.item_total,
                "tax_percentage": item.tax_percentage,
            }
            for index, item in enumerate(line_items)
        ],
        # Totals
        "totals": {
            "subtotal": po.sub_total,
            "tax_total": po.tax_total,
            "total": po.total,
            "discount": po.discount_amount,
            "adjustment": po.adjustment,
        },
        # Additional Details
        "notes": po.notes,
        "terms": po.terms,
        "custom_fields": po.custom_field_hash if po.custom_field_hash else {},
    }

    # Render the HTML template to string
    html_string = render_to_string("purchase_order_template.html", context)

    # Configure pdfkit options
    options = {
        "page-size": "A4",
        "margin-top": "0.35in",
        "margin-right": "0.35in",
        "margin-bottom": "0.35in",
        "margin-left": "0.35in",
        "encoding": "UTF-8",
        "no-outline": None,
        "enable-local-file-access": None,
    }

    # Generate PDF
    pdf = pdfkit.from_string(
        html_string, False, configuration=pdfkit_config, options=options
    )

    # Create the HTTP response
    response = HttpResponse(pdf, content_type="application/pdf")
    filename = f"PO_{po.purchaseorder_number}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


def format_currency(amount):
    """Format currency values"""
    if amount is None:
        return "0.00"
    return "{:,.2f}".format(float(amount))


def format_quantity(quantity):
    """Format quantity values"""
    if quantity is None:
        return "0"
    if float(quantity).is_integer():
        return str(int(float(quantity)))
    return "{:.2f}".format(float(quantity))


@api_view(["GET"])
def generate_bill_pdf(request, bill_id):
    """
    Generate PDF for a bill
    """
    try:
        entity = Entity.objects.last()
        # Get the bill instance
        bill = get_object_or_404(Bill, id=bill_id)

        # Get line items
        line_items = bill.bill_line_items.all().order_by("item_order")

        # Get vendor details from ZohoVendor if available
        vendor = bill.zoho_vendor

        # Prepare context data
        context = {
            "bill": {
                "bill_number": bill.bill_number,
                "reference_number": bill.reference_number,
                "date": bill.date,
                "due_date": bill.due_date,
                "payment_terms_label": bill.payment_terms_label,
                "notes": bill.notes,
                "terms": bill.terms,
                "sub_total": format_currency(bill.sub_total),
                "tax_total": format_currency(bill.tax_total),
                "total": format_currency(bill.total),
                "balance": format_currency(bill.balance),
                "currency_symbol": bill.currency_symbol or "₹",
                "custom_field_hash": bill.custom_field_hash or {},
            },
            "vendor": {
                "name": vendor.company_name if vendor else bill.vendor_name,
                "email": vendor.email if vendor else "",
                "phone": vendor.phone if vendor else "",
                "billing_address": bill.billing_address or {},
            },
            "line_items": [
                {
                    "description": item.description,
                    "quantity": format_quantity(item.quantity),
                    "rate": format_currency(item.rate),
                    "item_total": format_currency(item.item_total),
                }
                for item in line_items
            ],
            "company_info": {
                "name": entity.name,  # You might want to store this in settings
                "address_line": entity.billing_address,
            },
        }

        # Get the absolute path to your logo
        logo_path = "https://blueskycsr.com/wp-content/uploads/2023/10/Bluesky-Logo.png"
        context["logo_path"] = logo_path

        # Render the HTML template to string
        html_string = render_to_string("bill_pdf_template.html", context)

        # Configure pdfkit options
        options = {
            "page-size": "A4",
            "margin-top": "0.35in",
            "margin-right": "0.35in",
            "margin-bottom": "0.35in",
            "margin-left": "0.35in",
            "encoding": "UTF-8",
            "no-outline": None,
            "enable-local-file-access": None,
        }

        # Generate PDF
        pdf = pdfkit.from_string(
            html_string, False, configuration=pdfkit_config, options=options
        )

        # Create the HTTP response
        response = HttpResponse(pdf, content_type="application/pdf")
        filename = f"Bill_{bill.bill_number.replace('/', '_')}.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_vendor_login_data(request, vendor_id):
    vendor = Vendor.objects.get(vendor_id=vendor_id)
    response_data, response_status = get_role_response(vendor.user.user, "vendor", [])
    return Response(response_data, status=status.HTTP_200_OK)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def max_gmsheet_number(request):
    current_financial_year = get_current_financial_year_dates()
    try:
        # Get the latest gmsheet_number
        latest_gmsheet = GmSheet.objects.latest("created_at")
        gmsheet_number = latest_gmsheet.gmsheet_number

        # Extract the financial year part and the number part
        financial_year_part = gmsheet_number.split("/")[0]
        number_part = gmsheet_number.split("/")[-1][3:]

        # Ensure that the latest GmSheet is from the current financial year
        if financial_year_part == current_financial_year:
            latest_number = int(number_part)
            next_number = latest_number + 1
            next_gmsheet_number = f"{current_financial_year}/PRO{next_number:03d}"
        else:
            next_gmsheet_number = f"{current_financial_year}/PRO001"

    except GmSheet.DoesNotExist:
        next_gmsheet_number = f"{current_financial_year}/PRO001"

    return Response({"max_number": next_gmsheet_number})


class GMSheetDetailAPIView(APIView):
    def get(self, request, *args, **kwargs):
        gmsheet_id = request.GET.get("gmsheet_id")
        salesorder_ids = request.GET.getlist("salesorder_ids")

        if gmsheet_id:
            try:
                gm_sheet = GmSheet.objects.get(id=gmsheet_id)
                gm_sheet_data = GmSheetSerializer(gm_sheet).data
                return Response([gm_sheet_data], status=status.HTTP_200_OK)
            except GmSheet.DoesNotExist:
                return Response(
                    {"detail": "GM Sheet not found"}, status=status.HTTP_404_NOT_FOUND
                )
            except Exception as e:
                print(f"Error: {str(e)}")
                return Response(
                    {"detail": f"An error occurred: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        gm_sheets_data = []
        if salesorder_ids:
            try:
                salesorders = SalesOrder.objects.filter(
                    salesorder_id__in=salesorder_ids
                ).select_related("gm_sheet")

                for salesorder in salesorders:
                    if salesorder.gm_sheet:
                        gm_sheet_data = GmSheetSerializer(salesorder.gm_sheet).data
                        gm_sheet_data["salesorder_number"] = (
                            salesorder.salesorder_number
                        )
                        gm_sheets_data.append(gm_sheet_data)

            except Exception as e:
                print(f"Error: {str(e)}")
                return Response(
                    {"detail": f"An error occurred: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        if not gm_sheets_data:
            return Response(
                {"detail": "No GM Sheets found for the given criteria"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(gm_sheets_data, status=status.HTTP_200_OK)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_vendor(request):
    try:
        data = request.data
        with transaction.atomic():
            if Vendor.objects.filter(email=data["email"]).exists():
                return Response(
                    {"error": "Vendor with the same email already exists."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if all(
                key in data for key in ["name", "first_name", "email", "gst_treatment"]
            ):
                entity = Entity.objects.get(id=data.get("entity"))
                vendor_id = str(random.randint(1000000000, 9999999999))
                # Create ZohoVendor first
                zoho_vendor = ZohoVendor.objects.create(
                    contact_id=vendor_id,
                    contact_name=data.get("name"),
                    company_name=data.get("company_name", ""),
                    first_name=data.get("first_name"),
                    last_name=data.get("last_name", ""),
                    email=data.get("email"),
                    phone=data.get("phone", ""),
                    mobile=data.get(
                        "phone", ""
                    ),  # Using phone as mobile if mobile not provided
                    gst_treatment=data.get("gst_treatment"),
                    gst_no=data.get("gstn_uni", ""),
                    pan_no=data.get("pan", ""),
                    place_of_contact=data.get("place_of_contact", ""),
                    currency_id=data.get("currency", ""),
                    currency_code=data.get("currency", ""),
                    currency_symbol=get_currency_symbol(data.get("currency", "")),
                    tds_tax_id=data.get("tds", ""),
                    # Billing Address
                    billing_address={
                        "attention": data.get("attention", ""),
                        "country": data.get("country", ""),
                        "address": data.get("address", ""),
                        "city": data.get("city", ""),
                        "state": data.get("state", ""),
                        "zip": data.get("zip_code", ""),
                    },
                    # Shipping Address
                    shipping_address={
                        "attention": data.get("shipping_attention", ""),
                        "country": data.get("shipping_country", ""),
                        "address": data.get("shipping_address", ""),
                        "city": data.get("shipping_city", ""),
                        "state": data.get("shipping_state", ""),
                        "zip": data.get("shipping_zip_code", ""),
                    },
                    # Contact persons array
                    contact_persons=[
                        {
                            "first_name": data.get("first_name"),
                            "last_name": data.get("last_name", ""),
                            "email": data.get("email"),
                            "phone": data.get("phone", ""),
                        }
                    ],
                    # Bank account information (stored in bank_accounts JSON field)
                    bank_accounts=(
                        [
                            {
                                "beneficiary_name": data.get("beneficiary_name", ""),
                                "bank_name": data.get("bank_name", ""),
                                "account_number": data.get("account_number", ""),
                                "ifsc": data.get("ifsc", ""),
                            }
                        ]
                        if data.get("account_number")
                        else []
                    ),
                    # Additional fields
                    notes=data.get("remarks", ""),
                    status="active",
                    contact_type="vendor",
                    is_portal_enabled=True,
                    entity=entity,
                )

                # Handle user creation/retrieval
                user_profile = Profile.objects.filter(
                    user__username=data["email"]
                ).first()
                if user_profile:
                    user = user_profile.user
                else:
                    temp_password = "".join(
                        random.choices(string.ascii_letters + string.digits, k=8)
                    )
                    user = User.objects.create_user(
                        username=data["email"],
                        password=temp_password,
                        email=data["email"],
                        first_name=data.get("first_name", ""),
                        last_name=data.get("last_name", ""),
                    )
                    user_profile = Profile.objects.create(user=user)

                # Add vendor role
                vendor_role, created = Role.objects.get_or_create(name="vendor")
                user_profile.roles.add(vendor_role)

                # Create Vendor instance and link it to ZohoVendor
                vendor = Vendor.objects.create(
                    user=user_profile,
                    name=data["name"],
                    email=data["email"],
                    phone=data.get("phone", ""),
                    vendor_id=vendor_id,
                    zoho_vendor=zoho_vendor,
                )

                return Response(
                    {
                        "message": "Vendor created successfully!",
                        "vendor_id": vendor.id,
                        "zoho_vendor_id": zoho_vendor.id,
                    },
                    status=status.HTTP_201_CREATED,
                )

            else:
                return Response(
                    {"error": "Fill in all the required details."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
    except Exception as e:
        print(f"Error creating vendor: {str(e)}")
        return Response(
            {"error": "Failed to create vendor", "details": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
